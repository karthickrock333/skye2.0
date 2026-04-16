#!/usr/bin/env python3
"""
UAT Excel Test Runner & Evaluator
===================================
Reads the SKYE2.0 Feedback-UAT Excel spreadsheet, runs every question from the
Feedback Sheet against the live Skye agent with the correct tester metadata,
evaluates responses, and writes a new Excel file with:
  - All original columns preserved
  - "New Response (Post-Fix)"       — the agent's new response
  - "New Sources"                   — sources returned
  - "New Source Links"              — ServiceNow/GCS links
  - "Fix Applied"                   — description of what was fixed for this question
  - "New Assessment"                — auto-evaluated correctness vs tester feedback
  - "LLM Assessment"                — Gemini-based evaluation vs ideal response

Usage:
    python scripts/run_uat_excel.py                    # run all 152 questions
    python scripts/run_uat_excel.py --dry-run          # parse only, show plan
    python scripts/run_uat_excel.py --limit 10         # run first 10 only
    python scripts/run_uat_excel.py --start 50         # start from row 50
    python scripts/run_uat_excel.py --parallel 2       # 2 concurrent requests
    python scripts/run_uat_excel.py --reeval FILE.xlsx # re-evaluate existing results
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date
from pathlib import Path
from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests

# ── Config ──────────────────────────────────────────────────────────────────
AGENT_URL = os.getenv("SKYE_AGENT_URL", "http://localhost:8391")
CHAT_ENDPOINT = f"{AGENT_URL}/chat"
NEW_CHAT_ENDPOINT = f"{AGENT_URL}/new-chat"

# ── Auth for remote Cloud Run endpoints ────────────────────────────────────
# Set SKYE_AUTH_TOKEN for manual token, or leave unset to auto-fetch via ADC.
# Auth is only used when AGENT_URL points to a non-localhost endpoint.
_AUTH_TOKEN_ENV = os.getenv("SKYE_AUTH_TOKEN", "")
_auth_lock = threading.Lock()
_cached_id_token: str | None = None
_token_expiry: float = 0.0


def _is_remote_url() -> bool:
    """Return True if the agent URL is a remote (non-localhost) endpoint."""
    return not any(h in AGENT_URL for h in ("localhost", "127.0.0.1", "0.0.0.0"))


def _fetch_id_token_via_adc() -> tuple[str, float]:
    """Fetch a Google Cloud ID token using Application Default Credentials.

    Returns (token_string, expiry_timestamp).
    """
    try:
        import google.auth
        import google.auth.transport.requests
        from google.oauth2 import id_token as id_token_lib

        request = google.auth.transport.requests.Request()
        token = id_token_lib.fetch_id_token(request, AGENT_URL)
        # ID tokens are valid for 1 hour; refresh 5 min early
        expiry = time.time() + 3300
        return token, expiry
    except Exception as e:
        print(f"WARNING: Could not auto-fetch ID token via ADC: {e}")
        print("  Set SKYE_AUTH_TOKEN env var manually if needed.")
        return "", 0.0


def get_auth_headers() -> dict:
    """Return Authorization headers for the agent endpoint.

    - If SKYE_AUTH_TOKEN is set, use it directly (no auto-refresh).
    - If the URL is remote and no manual token, auto-fetch via ADC.
    - If the URL is localhost, return empty headers.
    """
    global _cached_id_token, _token_expiry

    if not _is_remote_url():
        return {}

    if _AUTH_TOKEN_ENV:
        return {"Authorization": f"Bearer {_AUTH_TOKEN_ENV}"}

    with _auth_lock:
        if _cached_id_token and time.time() < _token_expiry:
            return {"Authorization": f"Bearer {_cached_id_token}"}
        print("Fetching ID token via ADC...")
        _cached_id_token, _token_expiry = _fetch_id_token_via_adc()
        if _cached_id_token:
            return {"Authorization": f"Bearer {_cached_id_token}"}
    return {}


INPUT_FILE = Path(__file__).parent.parent / "docs" / "SKYE2.0 Feedback-UAT_Apr3.xlsx"
OUTPUT_DIR = Path(__file__).parent.parent / "test-results"

# ── Tester → Metadata Mapping ──────────────────────────────────────────────
# Maps tester names (from the Feedback Sheet) to their teams_metadata.
# Built from the Testers sheet + knowledge of their actual countries.
TESTER_METADATA = {
    "nikil jain": {
        "email": "nikil.jain@hitachidigital.com",
        "Country": "India",
        "usageLocation": "IN",
    },
    "dorota pajor": {
        "email": "dorota.pajor@hitachidigital.com",
        "Country": "Poland",
        "usageLocation": "PL",
        "officeLocation": "PL Warsaw",
    },
    "aparnaa": {
        "email": "aparnaa.udhayanan@hitachidigital.com",
        "Country": "India",
        "usageLocation": "IN",
    },
    "reema tiwari": {
        "email": "reema.tiwari@hitachidigital.com",
        "Country": "India",
        "usageLocation": "IN",
    },
    "priya chakraborty": {
        "email": "priya.chakraborty@hitachidigital.com",
        "Country": "India",
        "usageLocation": "IN",
    },
    "pavani gunnamanidi": {
        "email": "pavani.gunnamanidi@hitachidigital.com",
        "Country": "India",
        "usageLocation": "IN",
    },
    "boddhayan bhowmick": {
        "email": "boddhayan.bhowmick@hitachidigital.com",
        "Country": "India",
        "usageLocation": "IN",
    },
    "weronika wołek": {
        "email": "weronika.wolek@hitachidigital.com",
        "Country": "Poland",
        "usageLocation": "PL",
        "officeLocation": "PL Warsaw",
    },
    "angela picinic": {
        "email": "angela.picinic@hitachidigital.com",
        "Country": "US",
        "usageLocation": "US",
    },
    "daniela camacho": {
        "email": "daniela.camacho@hitachidigital.com",
        "Country": "US",
        "usageLocation": "US",
    },
    "krzysztof sowa": {
        "email": "krzysztof.sowa@hitachidigital.com",
        "Country": "Poland",
        "usageLocation": "PL",
    },
}

# Testers who are HR/GPS and should have global data_scope
HR_GPS_TESTERS = {
    "boddhayan bhowmick",
    "angela picinic",
    "dorota pajor",
    "aparnaa",
    "reema tiwari",
    "krzysztof sowa",
}

# ── Fix Descriptions ───────────────────────────────────────────────────────
# Maps issue patterns to fix descriptions for the "Fix Applied" column.
CROSS_COUNTRY_KB_FIX = (
    "Bug 7 fix: Cross-country KB contamination. Added KB→country override map "
    "(20 KB articles), system-name detection (Peoplepay/Darwin/Sodexo), and "
    "enhanced source filtering in post_validation_agent.py."
)
POLAND_FIX = (
    "Bug 9 fix: Poland holiday multi-layer fix. Fixed teams_metadata key casing "
    "(Country vs country), same-broad-region filtering in reranking (EMEA countries "
    "no longer cross-contaminate), KB source preference over legacy PDFs, "
    "region-augmented refined retrieval, and extended fallback detection."
)
ACCESS_CONTROL_FIX = (
    "Bug 9a fix: teams_metadata key casing — 'Country' (uppercase) now correctly "
    "read from Teams metadata. Also fixed for 'usageLocation' variants."
)
RERANKING_FIX = (
    "Bug 9b+9d fix: Same-broad-region cross-country filtering in reranking. "
    "UK/Spain/Italy/France articles no longer classified as 'regional' for Poland. "
    "ServiceNow KB articles preferred over legacy non-KB PDFs."
)
RETRIEVAL_FIX = (
    "Bug 9e fix: Refined holiday retrieval now includes target_region in the "
    "search query (e.g., 'holidays in Poland') for better vector search results."
)
FALLBACK_FIX = (
    "Bug 9f fix: Extended fallback phrase detection to catch 'I don't have specific' "
    "and 'I don't have a specific list' variations."
)
NO_FIX_NEEDED = "No fix needed — original response was correct."
PARTIAL_SOURCE_FIX = (
    "Bug 7+9 fixes: Cross-country source contamination resolved. Wrong-country "
    "KB sources now filtered by KB override map + metadata + region detection."
)

# ── LLM Evaluation ─────────────────────────────────────────────────────────
_llm_model = None
_llm_lock = threading.Lock()


def _get_llm():
    """Lazy-init Gemini model for LLM-as-judge evaluation."""
    global _llm_model
    if _llm_model is None:
        with _llm_lock:
            if _llm_model is None:
                from dotenv import load_dotenv

                load_dotenv()
                import vertexai
                from vertexai.generative_models import GenerativeModel

                project = os.getenv(
                    "EVAL_PROJECT_ID",
                    os.getenv("PROJECT_ID", "hd-procurement-poc-gemini"),
                )
                region = os.getenv("EVAL_REGION", os.getenv("REGION", "us-central1"))
                # Use separate credentials for evaluation if available
                eval_creds_path = os.getenv("EVAL_CREDENTIALS_PATH")
                if eval_creds_path:
                    from google.oauth2 import service_account

                    creds = service_account.Credentials.from_service_account_file(
                        eval_creds_path
                    )
                    vertexai.init(project=project, location=region, credentials=creds)
                else:
                    vertexai.init(project=project, location=region)
                _llm_model = GenerativeModel("gemini-2.0-flash")
    return _llm_model


_LLM_JUDGE_PROMPT = """\
You are an expert QA evaluator for an HR chatbot called SKYE. Your job is to \
judge whether the chatbot's NEW response correctly answers the user's question.

## Context
- **Today's Date**: {today_date}
- **User's Country**: {country}
- **User's Data Scope**: {data_scope}
- **Question**: {question}
- **Original Tester Accuracy**: {original_accuracy}
- **Tester Comments**: {tester_comments}
{ideal_section}
- **NEW Agent Response** (post-fix):
{new_response}

- **NEW Sources returned**: {new_sources}

## Important: Date Awareness
The UAT feedback was originally collected on April 3, 2026. Today's date is \
{today_date}. When evaluating time-sensitive answers (e.g. "next public \
holiday", "upcoming deadline"), use TODAY'S date, not the original test date. \
If the tester's ideal response mentions a date or event that has already passed \
(e.g. "Easter Monday" when today is after Easter), the agent's answer reflecting \
the NEXT upcoming date/event is CORRECT — do NOT penalize it for differing from \
the now-stale ideal. The agent should answer based on today's date.

## Important: Data Scope and Cross-Country Queries
{data_scope_instructions}

## Evaluation Criteria
1. **Correctness**: Does the response answer the question accurately?
2. **Country relevance**: Is the response about the correct country? See Data \
Scope rules above — global-scope users may legitimately ask about other countries.
3. **Completeness**: Does the response cover what the tester expected / the ideal \
response specifies?
4. **Source quality**: Are relevant KB sources provided? Do they link to the right \
country's policies?
5. **Regression check**: If the original was marked "Correct", is the new response \
at least as good?

## Comparison with Ideal Response
{ideal_instructions}

## Output Format
Respond with EXACTLY this JSON (no markdown fences, no extra text):
{{"verdict": "<VERDICT>", "score": <1-5>, "explanation": "<1-2 sentences>"}}

Where VERDICT is one of:
- CORRECT — response fully answers the question with correct country-specific info
- PARTIALLY_CORRECT — response addresses the question but is missing key info or \
has minor issues
- INCORRECT — response is wrong, about the wrong country, or critically incomplete
- REGRESSION — original was correct but the new response is worse
- NO_INFO — agent could not answer at all (returned "I don't have information")
"""


def _llm_evaluate(
    question: str,
    country: str,
    original_accuracy: str,
    tester_comments: str,
    ideal_response: str | None,
    new_response: str,
    new_sources: list[str],
    data_scope: str = "regional",
) -> dict:
    """Use Gemini to evaluate the new response quality.

    Returns: {"verdict": str, "score": int, "explanation": str}
    """
    # Build ideal response section
    if ideal_response and str(ideal_response).strip():
        ideal_section = f"- **Ideal Response** (from tester):\n{ideal_response}"
        ideal_instructions = (
            "The tester provided an Ideal Response. Compare the agent's new response "
            "against it. The new response does NOT need to match word-for-word, but it "
            "must convey the same key information and actionable guidance. If the ideal "
            "response mentions specific tools, links, or steps, check that the new "
            "response covers them."
        )
    else:
        ideal_section = "- **Ideal Response**: (not provided by tester)"
        ideal_instructions = (
            "No ideal response was provided. Evaluate based on the tester's comments, "
            "the original accuracy assessment, and general correctness for the "
            f"employee's country ({country}). Focus on whether the response is helpful, "
            "accurate, and country-appropriate."
        )

    # Build data_scope instructions
    if data_scope == "global":
        data_scope_instructions = (
            f"This user has **global** data access (they are an HR/GPS professional "
            f"based in {country}). They can legitimately ask about ANY country's "
            f"policies, not just {country}. If the question mentions another country "
            f"(e.g., 'US benefits', '401(k) info', 'employment verification in the US'), "
            f"the response SHOULD answer about that specific country — do NOT penalize "
            f"the response for answering about a different country when the user "
            f"explicitly asked about it. Only penalize if the response returns info for "
            f"a country the user did NOT ask about."
        )
    else:
        data_scope_instructions = (
            f"This user has **regional** data access (based in {country}). "
            f"Sources from other countries are a critical failure unless the user "
            f"explicitly asked about another country."
        )

    prompt = _LLM_JUDGE_PROMPT.format(
        today_date=date.today().strftime("%A, %B %d, %Y"),
        country=country,
        data_scope=data_scope,
        question=question,
        original_accuracy=original_accuracy or "(not assessed)",
        tester_comments=tester_comments or "(no comments)",
        ideal_section=ideal_section,
        ideal_instructions=ideal_instructions,
        data_scope_instructions=data_scope_instructions,
        new_response=(new_response or "(empty)")[:3000],
        new_sources=", ".join(new_sources[:10]) if new_sources else "(none)",
    )

    try:
        model = _get_llm()
        resp = model.generate_content(prompt)
        text = resp.text.strip()
        # Strip markdown fences if present
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
        result = json.loads(text)
        # Validate
        if "verdict" not in result:
            result["verdict"] = "UNKNOWN"
        if "score" not in result:
            result["score"] = 0
        if "explanation" not in result:
            result["explanation"] = ""
        return result
    except json.JSONDecodeError as e:
        return {
            "verdict": "PARSE_ERROR",
            "score": 0,
            "explanation": f"LLM output not valid JSON: {e}",
        }
    except Exception as e:
        return {"verdict": "ERROR", "score": 0, "explanation": str(e)[:200]}


def _determine_fix_description(
    original_accuracy: str,
    original_comments: str,
    question: str,
    country: str,
    new_sources: list,
) -> str:
    """Determine what fix was applied based on tester feedback patterns."""
    acc = (original_accuracy or "").lower().strip()
    comments = (original_comments or "").lower()

    # Correct responses — no fix needed
    if "correct" in acc and "partial" not in acc and "incorrect" not in acc:
        return NO_FIX_NEEDED

    # Cross-country KB contamination patterns
    kb_patterns = [
        "incorrect kb",
        "wrong kb",
        "incorrect source",
        "irrelevant source",
        "wrong source",
        "redirects to incorrect",
        "india kb",
        "uk kb",
        "indonesia",
        "vietnam",
        "malaysia",
        "spain",
        "france",
        "austria",
        "argentina",
        "thailand",
        "israel",
        "other country",
        "source referenced",
        "taking me to uk",
        "taking me to",
    ]
    if any(p in comments for p in kb_patterns):
        if country == "Poland":
            return POLAND_FIX + " " + CROSS_COUNTRY_KB_FIX
        return CROSS_COUNTRY_KB_FIX

    # Poland-specific issues
    if country == "Poland":
        if "holiday" in question.lower() or "public holiday" in question.lower():
            return POLAND_FIX
        if "incorrect" in comments or "wrong" in comments:
            return POLAND_FIX + " " + CROSS_COUNTRY_KB_FIX

    # Access control issues (HR should have global access)
    if "global access" in comments or "hrs should" in comments:
        return ACCESS_CONTROL_FIX

    # Location constraint / wrong country response
    if "location constraint" in comments:
        return ACCESS_CONTROL_FIX

    # Source issues without specific country mention
    if "source" in comments and ("incorrect" in comments or "irrelevant" in comments):
        return PARTIAL_SOURCE_FIX

    # No KB / no source provided
    if "no source" in comments or "no kb" in comments:
        return PARTIAL_SOURCE_FIX

    # Generic partial/incorrect without specific pattern
    if "incorrect" in acc or "partial" in acc:
        return (
            "General improvements: enhanced retrieval, reranking, and source filtering."
        )

    return ""


def _auto_evaluate(
    original_accuracy: str,
    original_comments: str,
    new_answer: str,
    new_sources: list,
    question: str,
    country: str,
    ideal_response: str,
) -> str:
    """Auto-evaluate the new response vs tester expectations."""
    acc = (original_accuracy or "").lower().strip()
    comments = (original_comments or "").lower()
    answer_lower = new_answer.lower()

    # If original was correct, check if still correct
    if "correct" in acc and "partial" not in acc and "incorrect" not in acc:
        if "i don't have" in answer_lower and "information" in answer_lower:
            return "REGRESSION — was correct, now returning no-info"
        return "STILL CORRECT"

    # Check for cross-country KB contamination
    source_str = " ".join(new_sources).lower()
    cross_country_keywords = {
        "india": [
            "uk",
            "indonesia",
            "vietnam",
            "malaysia",
            "spain",
            "france",
            "austria",
            "argentina",
            "thailand",
            "israel",
        ],
        "poland": [
            "india",
            "indonesia",
            "vietnam",
            "malaysia",
            "argentina",
            "thailand",
        ],
    }
    bad_countries = cross_country_keywords.get(country.lower(), [])
    has_cross_country = any(bc in source_str for bc in bad_countries)

    if has_cross_country:
        return "STILL HAS CROSS-COUNTRY SOURCE"

    # Check if tester said "incorrect source" and we now have better sources
    if (
        "incorrect source" in comments
        or "wrong kb" in comments
        or "incorrect kb" in comments
    ):
        if new_sources:
            return "IMPROVED — sources changed (verify manually)"
        return "PARTIAL — no sources returned"

    # Check for no-info responses
    no_info_patterns = [
        "i don't have information",
        "i don't have specific",
        "i don't have enough",
        "no relevant context",
    ]
    if any(p in answer_lower for p in no_info_patterns):
        return "NO INFO — agent couldn't answer"

    # If original was incorrect and we got an answer with sources
    if "incorrect" in acc:
        if new_sources:
            return "IMPROVED — has answer + sources (verify manually)"
        return "CHANGED — verify manually"

    # Partial responses
    if "partial" in acc:
        if new_sources:
            return "IMPROVED — has answer + sources (verify manually)"
        return "CHANGED — verify manually"

    return "NEEDS MANUAL REVIEW"


def _resolve_tester_metadata(tester_name: str, country: str) -> dict:
    """Resolve teams_metadata for a tester based on name and question country."""
    if not tester_name:
        return {"email": "test.user@hitachidigital.com", "Country": country or "India"}

    tester_lower = tester_name.strip().lower()
    # Try exact match
    meta = TESTER_METADATA.get(tester_lower)
    if meta:
        return dict(meta)

    # Try partial match
    for key, val in TESTER_METADATA.items():
        if key in tester_lower or tester_lower in key:
            return dict(val)

    # Fallback: construct from available info
    email_name = tester_lower.replace(" ", ".").replace("ł", "l").replace("ó", "o")
    return {
        "email": f"{email_name}@hitachidigital.com",
        "Country": country or "India",
        "usageLocation": "IN" if (country or "").lower() == "india" else "PL",
    }


def _resolve_data_scope(tester_name: str) -> str:
    """Determine data_scope based on tester role."""
    tester_lower = (tester_name or "").strip().lower()
    if tester_lower in HR_GPS_TESTERS:
        return "global"
    return "regional"


def send_question(
    question: str,
    session_id: str,
    teams_metadata: dict,
    data_scope: str = "regional",
    timeout: int = 120,
) -> dict:
    """Send a question to the Skye agent and return the response."""
    payload = {
        "question": question,
        "session_id": session_id,
        "teams_metadata": teams_metadata,
        "data_scope": data_scope,
    }
    try:
        resp = requests.post(
            CHAT_ENDPOINT, json=payload, headers=get_auth_headers(), timeout=timeout
        )
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.Timeout:
        return {"answer": "[TIMEOUT]", "sources": [], "source_links": {}}
    except Exception as e:
        return {"answer": f"[ERROR: {e}]", "sources": [], "source_links": {}}


def reset_session(session_id: str):
    """Reset agent session to avoid conversation history cross-contamination."""
    try:
        requests.post(
            NEW_CHAT_ENDPOINT,
            json={"session_id": session_id},
            headers=get_auth_headers(),
            timeout=10,
        )
    except Exception:
        pass


def run_uat_tests(
    input_file: Path,
    output_dir: Path,
    dry_run: bool = False,
    limit: int | None = None,
    start: int = 1,
    parallel: int = 1,
) -> Path:
    """Run UAT tests and generate output Excel."""
    wb = openpyxl.load_workbook(str(input_file))
    ws = wb["Feedback Sheet"]

    # ── Discover original columns ────────────────────────────────────────
    orig_headers = []
    for c in range(1, ws.max_column + 1):
        orig_headers.append(ws.cell(1, c).value)
    print(f"Original columns ({len(orig_headers)}): {orig_headers}")

    # ── Add new columns ──────────────────────────────────────────────────
    new_cols = [
        "New Response (Post-Fix)",
        "New Sources",
        "New Source Links",
        "Fix Applied",
        "New Assessment",
        "Response Time (s)",
    ]
    start_col = ws.max_column + 1
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for i, col_name in enumerate(new_cols):
        cell = ws.cell(1, start_col + i, col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Set column widths for new columns
    for i, col_name in enumerate(new_cols):
        col_letter = openpyxl.utils.get_column_letter(start_col + i)
        if "Response" in col_name and "Time" not in col_name:
            ws.column_dimensions[col_letter].width = 60
        elif "Fix" in col_name:
            ws.column_dimensions[col_letter].width = 50
        elif "Assessment" in col_name:
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 35

    # ── Collect questions ────────────────────────────────────────────────
    questions = []
    for r in range(2, ws.max_row + 1):
        sr_no = ws.cell(r, 1).value
        question = ws.cell(r, 5).value
        if not question or not str(question).strip():
            continue
        questions.append(
            {
                "row": r,
                "sr_no": sr_no,
                "tester": ws.cell(r, 3).value,
                "country": ws.cell(r, 4).value or "India",
                "question": str(question).strip(),
                "original_response": ws.cell(r, 6).value,
                "score": ws.cell(r, 7).value,
                "accuracy": ws.cell(r, 8).value,
                "comments": ws.cell(r, 9).value,
                "ideal": ws.cell(r, 10).value,
                "team_comments": ws.cell(r, 11).value,
            }
        )

    total = len(questions)
    # Apply start/limit filters
    filtered = [q for q in questions if q["sr_no"] and int(q["sr_no"]) >= start]
    if limit:
        filtered = filtered[:limit]

    print(f"\nTotal questions: {total}")
    print(f"Running: {len(filtered)} (start={start}, limit={limit})")

    if dry_run:
        print("\n=== DRY RUN — Questions to be tested ===")
        for q in filtered:
            meta = _resolve_tester_metadata(q["tester"], q["country"])
            scope = _resolve_data_scope(q["tester"])
            print(
                f"  #{q['sr_no']} [{q['country']}] [{scope}] "
                f"{q['tester']}: {q['question'][:80]}"
            )
        return None

    # ── Run tests ────────────────────────────────────────────────────────
    print(f"\nRunning {len(filtered)} questions against {AGENT_URL}...")
    print(f"Parallel workers: {parallel}")
    print("=" * 70)

    results = {}
    total_start = time.time()

    def _run_one(q: dict) -> tuple:
        sr = q["sr_no"]
        session_id = f"uat_retest_{sr}_{int(time.time())}"
        meta = _resolve_tester_metadata(q["tester"], q["country"])
        scope = _resolve_data_scope(q["tester"])

        # Reset session to avoid history contamination
        reset_session(session_id)

        t0 = time.time()
        resp = send_question(q["question"], session_id, meta, scope)
        elapsed = time.time() - t0

        return sr, resp, elapsed

    if parallel > 1:
        with ThreadPoolExecutor(max_workers=parallel) as pool:
            futures = {pool.submit(_run_one, q): q for q in filtered}
            done = 0
            for fut in as_completed(futures):
                q = futures[fut]
                sr, resp, elapsed = fut.result()
                results[sr] = (resp, elapsed)
                done += 1
                print(
                    f"  [{done}/{len(filtered)}] #{sr} ({elapsed:.1f}s) "
                    f"{q['question'][:50]}... → "
                    f"{len(resp.get('sources', []))} sources"
                )
    else:
        for i, q in enumerate(filtered):
            sr, resp, elapsed = _run_one(q)
            results[sr] = (resp, elapsed)
            answer_preview = resp.get("answer", "")[:60].replace("\n", " ")
            print(
                f"  [{i + 1}/{len(filtered)}] #{sr} ({elapsed:.1f}s) "
                f"{q['question'][:50]}... → "
                f"{len(resp.get('sources', []))} src | {answer_preview}..."
            )

    total_time = time.time() - total_start
    print(f"\n{'=' * 70}")
    print(f"Completed {len(results)} questions in {total_time:.1f}s")

    # ── Write results back to Excel ──────────────────────────────────────
    # Color fills for assessment
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    improved_count = 0
    still_correct_count = 0
    regression_count = 0
    needs_review_count = 0

    for q in filtered:
        sr = q["sr_no"]
        row = q["row"]
        if sr not in results:
            continue

        resp, elapsed = results[sr]
        answer = resp.get("answer", "")
        sources = resp.get("sources", [])
        source_links = resp.get("source_links", {})

        # Determine fix description
        fix_desc = _determine_fix_description(
            q["accuracy"], q["comments"], q["question"], q["country"], sources
        )

        # Auto-evaluate (heuristic)
        assessment = _auto_evaluate(
            q["accuracy"],
            q["comments"],
            answer,
            sources,
            q["question"],
            q["country"],
            q.get("ideal"),
        )

        # Write new columns
        # Col: New Response
        cell = ws.cell(row, start_col, answer)
        cell.alignment = wrap_align

        # Col: New Sources
        cell = ws.cell(row, start_col + 1, "\n".join(sources) if sources else "(none)")
        cell.alignment = wrap_align

        # Col: New Source Links
        links_str = "\n".join(f"{k}: {v}" for k, v in source_links.items())
        cell = ws.cell(row, start_col + 2, links_str if links_str else "(none)")
        cell.alignment = wrap_align

        # Col: Fix Applied
        cell = ws.cell(row, start_col + 3, fix_desc)
        cell.alignment = wrap_align

        # Col: New Assessment (heuristic)
        cell = ws.cell(row, start_col + 4, assessment)
        cell.alignment = wrap_align
        if "STILL CORRECT" in assessment or "IMPROVED" in assessment:
            cell.fill = green_fill
            if "IMPROVED" in assessment:
                improved_count += 1
            else:
                still_correct_count += 1
        elif "REGRESSION" in assessment or "CROSS-COUNTRY" in assessment:
            cell.fill = red_fill
            regression_count += 1
        else:
            cell.fill = yellow_fill
            needs_review_count += 1

        # Col: Response Time
        cell = ws.cell(row, start_col + 5, round(elapsed, 1))
        cell.alignment = Alignment(horizontal="center")

    # ── Save output ──────────────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"UAT_Results_{timestamp}.xlsx"
    wb.save(str(output_file))

    # ── Print summary ────────────────────────────────────────────────────
    print(f"\n{'=' * 70}")
    print(f"RESULTS SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Total tested:     {len(results)}")
    print(f"  Still correct:    {still_correct_count}")
    print(f"  Improved:         {improved_count}")
    print(f"  Regressions:      {regression_count}")
    print(f"  Needs review:     {needs_review_count}")
    print(f"  Total time:       {total_time:.1f}s")
    print(f"  Avg per question: {total_time / len(results):.1f}s")
    print(f"\n  Output: {output_file}")

    return output_file


# ── Re-evaluation mode ──────────────────────────────────────────────────────
def reeval_results(
    input_file: Path,
    output_dir: Path,
    limit: int | None = None,
    start: int = 1,
    parallel: int = 3,
) -> Path:
    """Re-evaluate an existing results Excel using LLM-as-judge.

    Reads the results file (which already has "New Response (Post-Fix)" etc.),
    runs Gemini evaluation on each row, and adds/updates these columns:
      - "LLM Verdict"     — CORRECT / PARTIALLY_CORRECT / INCORRECT / etc.
      - "LLM Score"       — 1-5 quality score
      - "LLM Explanation" — Brief explanation of the verdict

    This does NOT re-query the agent — it only evaluates existing responses.
    """
    wb = openpyxl.load_workbook(str(input_file))
    ws = wb["Feedback Sheet"]

    # ── Discover columns ─────────────────────────────────────────────────
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[h] = c
    print(f"Columns found: {list(headers.keys())}")

    # Required columns from original + previous run
    col_sr = headers.get("Sr.No.", 1)
    col_tester = headers.get("Tester Email ID/ Name", 3)
    col_country = headers.get("Country", 4)
    col_question = headers.get("Question", 5)
    col_orig_accuracy = headers.get("Response Accuracy Assessment", 8)
    col_comments = headers.get("Tester Comments", 9)
    col_ideal = headers.get("Ideal Response (If any)", 10)
    col_new_response = headers.get("New Response (Post-Fix)")
    col_new_sources = headers.get("New Sources")
    col_new_assessment = headers.get("New Assessment")

    if not col_new_response:
        print(
            "ERROR: 'New Response (Post-Fix)' column not found — is this a results file?"
        )
        sys.exit(1)

    # ── Add LLM columns (or find existing ones) ─────────────────────────
    llm_cols = ["LLM Verdict", "LLM Score", "LLM Explanation"]
    llm_start_col = None
    for lc_name in llm_cols:
        if lc_name in headers:
            if llm_start_col is None:
                llm_start_col = headers[lc_name]
            break

    if llm_start_col is None:
        llm_start_col = ws.max_column + 1
        header_fill = PatternFill(
            start_color="7030A0", end_color="7030A0", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for i, col_name in enumerate(llm_cols):
            cell = ws.cell(1, llm_start_col + i, col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Set column widths
        for i, col_name in enumerate(llm_cols):
            col_letter = openpyxl.utils.get_column_letter(llm_start_col + i)
            if "Explanation" in col_name:
                ws.column_dimensions[col_letter].width = 60
            elif "Verdict" in col_name:
                ws.column_dimensions[col_letter].width = 22
            else:
                ws.column_dimensions[col_letter].width = 12

    col_llm_verdict = llm_start_col
    col_llm_score = llm_start_col + 1
    col_llm_explanation = llm_start_col + 2

    # ── Collect rows to evaluate ─────────────────────────────────────────
    rows_to_eval = []
    for r in range(2, ws.max_row + 1):
        sr_no = ws.cell(r, col_sr).value
        question = ws.cell(r, col_question).value
        new_response = ws.cell(r, col_new_response).value
        if not question or not new_response:
            continue
        if sr_no and int(sr_no) < start:
            continue
        rows_to_eval.append(
            {
                "row": r,
                "sr_no": sr_no,
                "tester": ws.cell(r, col_tester).value,
                "country": ws.cell(r, col_country).value or "India",
                "question": str(question).strip(),
                "original_accuracy": ws.cell(r, col_orig_accuracy).value,
                "comments": ws.cell(r, col_comments).value,
                "ideal": ws.cell(r, col_ideal).value,
                "new_response": str(new_response).strip(),
                "new_sources": (ws.cell(r, col_new_sources).value or "").split("\n"),
                "heuristic_assessment": ws.cell(r, col_new_assessment).value,
            }
        )

    if limit:
        rows_to_eval = rows_to_eval[:limit]

    print(f"\nRows to evaluate with LLM: {len(rows_to_eval)}")
    print(f"Using Gemini 2.0 Flash as judge")
    print("=" * 70)

    # ── Run LLM evaluations ──────────────────────────────────────────────
    llm_results = {}
    total_start = time.time()

    def _eval_one(item: dict) -> tuple:
        sr = item["sr_no"]
        tester_name = (item.get("tester") or "").strip().lower()
        scope = "global" if tester_name in HR_GPS_TESTERS else "regional"
        result = _llm_evaluate(
            question=item["question"],
            country=item["country"],
            original_accuracy=item["original_accuracy"],
            tester_comments=item["comments"],
            ideal_response=item["ideal"],
            new_response=item["new_response"],
            new_sources=item["new_sources"],
            data_scope=scope,
        )
        return sr, item["row"], result

    done_count = 0
    if parallel > 1:
        with ThreadPoolExecutor(max_workers=parallel) as pool:
            futures = {pool.submit(_eval_one, item): item for item in rows_to_eval}
            for fut in as_completed(futures):
                item = futures[fut]
                try:
                    sr, row, result = fut.result()
                    llm_results[sr] = (row, result)
                    done_count += 1
                    print(
                        f"  [{done_count}/{len(rows_to_eval)}] #{sr} "
                        f"→ {result['verdict']} (score={result['score']}) "
                        f"| {result['explanation'][:60]}"
                    )
                except Exception as e:
                    done_count += 1
                    sr = item["sr_no"]
                    llm_results[sr] = (
                        item["row"],
                        {
                            "verdict": "ERROR",
                            "score": 0,
                            "explanation": str(e)[:200],
                        },
                    )
                    print(f"  [{done_count}/{len(rows_to_eval)}] #{sr} → ERROR: {e}")
    else:
        for item in rows_to_eval:
            sr, row, result = _eval_one(item)
            llm_results[sr] = (row, result)
            done_count += 1
            print(
                f"  [{done_count}/{len(rows_to_eval)}] #{sr} "
                f"→ {result['verdict']} (score={result['score']}) "
                f"| {result['explanation'][:60]}"
            )

    total_time = time.time() - total_start

    # ── Write LLM results to Excel ───────────────────────────────────────
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    verdict_counts = {}
    score_sum = 0
    score_count = 0

    for sr, (row, result) in llm_results.items():
        verdict = result.get("verdict", "UNKNOWN")
        score = result.get("score", 0)
        explanation = result.get("explanation", "")

        verdict_counts[verdict] = verdict_counts.get(verdict, 0) + 1
        if score > 0:
            score_sum += score
            score_count += 1

        # Write verdict
        cell = ws.cell(row, col_llm_verdict, verdict)
        cell.alignment = wrap_align
        if verdict == "CORRECT":
            cell.fill = green_fill
        elif verdict in ("INCORRECT", "REGRESSION"):
            cell.fill = red_fill
        elif verdict == "PARTIALLY_CORRECT":
            cell.fill = yellow_fill
        elif verdict == "NO_INFO":
            cell.fill = red_fill
        else:
            cell.fill = yellow_fill

        # Write score
        cell = ws.cell(row, col_llm_score, score)
        cell.alignment = Alignment(horizontal="center")
        if score >= 4:
            cell.fill = green_fill
        elif score >= 3:
            cell.fill = yellow_fill
        elif score > 0:
            cell.fill = red_fill

        # Write explanation
        cell = ws.cell(row, col_llm_explanation, explanation)
        cell.alignment = wrap_align

    # ── Also update the heuristic "New Assessment" using LLM verdicts ────
    # For rows where the heuristic said "verify manually", override with LLM
    if col_new_assessment:
        for sr, (row, result) in llm_results.items():
            current = ws.cell(row, col_new_assessment).value or ""
            verdict = result.get("verdict", "")
            score = result.get("score", 0)

            # Map LLM verdict to a cleaner heuristic label
            if verdict == "CORRECT":
                new_label = "IMPROVED — LLM verified correct"
            elif verdict == "PARTIALLY_CORRECT" and score >= 3:
                new_label = "IMPROVED — partially correct (LLM)"
            elif verdict == "INCORRECT":
                new_label = "NOT IMPROVED — LLM says incorrect"
            elif verdict == "REGRESSION":
                new_label = "REGRESSION — LLM confirmed"
            elif verdict == "NO_INFO":
                new_label = "NO INFO — agent couldn't answer"
            else:
                new_label = None  # Keep heuristic as-is

            if new_label and "verify manually" in current.lower():
                cell = ws.cell(row, col_new_assessment, new_label)
                cell.alignment = wrap_align
                if "IMPROVED" in new_label:
                    cell.fill = green_fill
                elif (
                    "REGRESSION" in new_label
                    or "NOT IMPROVED" in new_label
                    or "NO INFO" in new_label
                ):
                    cell.fill = red_fill
                else:
                    cell.fill = yellow_fill

    # ── Save ─────────────────────────────────────────────────────────────
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"UAT_Results_LLM_{timestamp}.xlsx"
    wb.save(str(output_file))

    # ── Print summary ────────────────────────────────────────────────────
    avg_score = score_sum / score_count if score_count else 0
    print(f"\n{'=' * 70}")
    print(f"LLM EVALUATION SUMMARY")
    print(f"{'=' * 70}")
    print(f"  Total evaluated:  {len(llm_results)}")
    print(f"  Average score:    {avg_score:.1f} / 5")
    print(f"  Eval time:        {total_time:.1f}s")
    print(f"  Avg per question: {total_time / max(len(llm_results), 1):.1f}s")
    print(f"\n  Verdict Distribution:")
    for v, c in sorted(verdict_counts.items(), key=lambda x: -x[1]):
        pct = c / len(llm_results) * 100
        print(f"    {v:25s} {c:3d}  ({pct:.0f}%)")
    print(f"\n  Output: {output_file}")

    return output_file


def main():
    parser = argparse.ArgumentParser(description="UAT Excel Test Runner & Evaluator")
    parser.add_argument("--dry-run", action="store_true", help="Parse only, don't run")
    parser.add_argument("--limit", type=int, default=None, help="Max questions to run")
    parser.add_argument("--start", type=int, default=1, help="Start from Sr.No.")
    parser.add_argument("--parallel", type=int, default=1, help="Concurrent requests")
    parser.add_argument("--input", type=str, default=None, help="Input Excel path")
    parser.add_argument(
        "--reeval",
        type=str,
        default=None,
        help="Re-evaluate an existing results Excel with LLM judge (no agent calls)",
    )
    args = parser.parse_args()

    # ── Re-evaluation mode ───────────────────────────────────────────────
    if args.reeval:
        reeval_file = Path(args.reeval)
        if not reeval_file.exists():
            print(f"ERROR: File not found: {reeval_file}")
            sys.exit(1)
        output = reeval_results(
            input_file=reeval_file,
            output_dir=OUTPUT_DIR,
            limit=args.limit,
            start=args.start,
            parallel=args.parallel,
        )
        if output:
            print(f"\nDone! Open the Excel file to review LLM evaluations.")
        return

    # ── Normal run mode ──────────────────────────────────────────────────
    input_file = Path(args.input) if args.input else INPUT_FILE
    if not input_file.exists():
        print(f"ERROR: Input file not found: {input_file}")
        sys.exit(1)

    # Health check — use /cache/status for remote (Cloud Run intercepts /healthz)
    if not args.dry_run:
        try:
            health_path = "/cache/status" if _is_remote_url() else "/healthz"
            resp = requests.get(
                f"{AGENT_URL}{health_path}",
                headers=get_auth_headers(),
                timeout=15,
            )
            resp.raise_for_status()
            print(f"Agent healthy at {AGENT_URL} (via {health_path})")
        except Exception as e:
            print(f"ERROR: Agent not reachable at {AGENT_URL}: {e}")
            sys.exit(1)

    output = run_uat_tests(
        input_file=input_file,
        output_dir=OUTPUT_DIR,
        dry_run=args.dry_run,
        limit=args.limit,
        start=args.start,
        parallel=args.parallel,
    )

    if output:
        print(f"\nDone! Open the Excel file to review results.")


if __name__ == "__main__":
    main()
