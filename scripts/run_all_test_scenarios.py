"""
run_all_test_scenarios.py
Reads test cases from SKYE_HR_Agent_Test_Cases.xlsx,
calls the SKYE chat API for each, and writes responses back.
Uses unique session_id per email to avoid cross-contamination.
"""

import time
import uuid
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ── Config ───────────────────────────────────────────────────────────────────
API_URL = "https://hd-hd1ai-skye-agent-j2igqzmkga-uc.a.run.app/chat"
TOKEN = (
    "eyJhbGciOiJSUzI1NiIsImtpZCI6ImIzZDk1Yjk1ZmE0OGQxODBiODVmZmU4MDgyZmNmYTIxNzRiMDQ2NjciLCJ0eXAiOiJKV1QifQ"
    ".eyJpc3MiOiJodHRwczovL2FjY291bnRzLmdvb2dsZS5jb20iLCJhenAiOiIzMjU1NTk0MDU1OS5hcHBzLmdvb2dsZXVzZXJjb250ZW50LmNvbSIsImF1ZCI6IjMyNTU1OTQwNTU5LmFwcHMuZ29vZ2xldXNlcmNvbnRlbnQuY29tIiwic3ViIjoiMTA1MDA0MDI0NTI5NTIyNDQ5MjA4IiwiaGQiOiJoaXRhY2hpZGlnaXRhbC5jb20iLCJlbWFpbCI6ImFudWoua2FybkBoaXRhY2hpZGlnaXRhbC5jb20iLCJlbWFpbF92ZXJpZmllZCI6dHJ1ZSwiYXRfaGFzaCI6IktsbHFTOW5CUDJPWUVUMEF1R0g0U2ciLCJpYXQiOjE3NzYyNDgxMTYsImV4cCI6MTc3NjI1MTcxNn0"
    ".Oxo_wX8JN0-QVWy6a7DVEH_AHfDchEpP_rvd4QYhYyjz19wM7GLutCcfemCglQP05PE3H6_SrQ_-l0Xgo4TYB9TmBMhbWw9qPxrECFKDzldYN6L1LpXmQndi3mTEBiGsL4Ry9LuBjsFovdXvvN93h5J0SOYvi-UDoxuxofRp0KzdrormtaJv00JcstUhd5idG5Y40EFQixIvR6N2PbTerERsIPXpq0AvpBjAijG75TeEyVUf55lFJKz3l_RiXJf7iAwRlfqswGrf0oQkl0GLZZxUWsbKcUBklwZHh0FSj4IME0Xn4clytfDcKMPN90syLPKWOrcM2z2oOuI6JIxxMw"
)

EXCEL_PATH = r"c:\Users\SSLTP11340\Desktop\SKYE\Testing Prompts\SKYE_HR_Agent_Test_Cases.xlsx"
HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Content-Type": "application/json",
}
TIMEOUT = 120  # seconds per request

# Column indices (1-based) in the Excel
COL_SNO = 1
COL_EMAIL = 5
COL_QUESTION = 7
COL_RESPONSE = 9
COL_PASS_FAIL = 10

WRAP = Alignment(wrap_text=True, vertical="top")


def generate_session_id(email: str, email_sessions: dict) -> str:
    """Generate a unique session_id per email. Reuse within same email."""
    if email not in email_sessions:
        short_id = uuid.uuid4().hex[:8]
        clean = email.split("@")[0].replace(".", "_")[:20]
        email_sessions[email] = f"test_{clean}_{short_id}"
    return email_sessions[email]


def call_api(question: str, email: str, session_id: str) -> tuple:
    """Call SKYE chat API. Returns (response_text, duration_seconds)."""
    payload = {
        "question": question,
        "session_id": session_id,
        "teams_metadata": {"email": email},
        "data_scope": "regional",
    }
    t0 = time.time()
    try:
        resp = requests.post(API_URL, json=payload, headers=HEADERS, timeout=TIMEOUT)
        elapsed = round(time.time() - t0, 1)
        if resp.status_code == 200:
            data = resp.json()
            # Try common response fields
            answer = (
                data.get("response")
                or data.get("answer")
                or data.get("message")
                or data.get("text")
                or str(data)
            )
            return answer, elapsed
        else:
            return f"HTTP {resp.status_code}: {resp.text[:300]}", elapsed
    except requests.exceptions.Timeout:
        return "TIMEOUT (>120s)", round(time.time() - t0, 1)
    except Exception as e:
        return f"ERROR: {str(e)[:300]}", round(time.time() - t0, 1)


def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    email_sessions = {}
    total = 0
    success = 0
    errors = 0

    # Count total test cases first
    test_rows = []
    for row_idx in range(2, ws.max_row + 1):
        sno = ws.cell(row=row_idx, column=COL_SNO).value
        email = ws.cell(row=row_idx, column=COL_EMAIL).value
        question = ws.cell(row=row_idx, column=COL_QUESTION).value
        if sno and email and question:
            test_rows.append(row_idx)

    print(f"Found {len(test_rows)} test cases to run")
    print("=" * 70)

    for i, row_idx in enumerate(test_rows):
        sno = ws.cell(row=row_idx, column=COL_SNO).value
        email = str(ws.cell(row=row_idx, column=COL_EMAIL).value).strip()
        question = str(ws.cell(row=row_idx, column=COL_QUESTION).value).strip()

        session_id = generate_session_id(email, email_sessions)
        total += 1

        print(f"[{total}/{len(test_rows)}] S.No {sno} | {email[:30]}...")
        print(f"  Q: {question[:80]}...")

        answer, elapsed = call_api(question, email, session_id)

        # Write response to Excel
        resp_cell = ws.cell(row=row_idx, column=COL_RESPONSE)
        resp_cell.value = str(answer)[:32000]  # Excel cell limit
        resp_cell.alignment = WRAP

        if answer.startswith("HTTP ") or answer.startswith("ERROR") or answer == "TIMEOUT (>120s)":
            errors += 1
            print(f"  !! {answer[:100]}")
            # Check if token expired
            if "401" in str(answer) or "403" in str(answer):
                print("\n!! TOKEN EXPIRED — saving progress and stopping.")
                print(f"   Completed {total} of {len(test_rows)} tests.")
                print(f"   Re-run script after refreshing the token.\n")
                wb.save(EXCEL_PATH)
                return
        else:
            success += 1
            preview = str(answer)[:120].replace("\n", " ")
            print(f"  A: {preview}... ({elapsed}s)")

        # Save every 10 requests (in case of crash)
        if total % 10 == 0:
            wb.save(EXCEL_PATH)
            print(f"  [checkpoint saved — {total}/{len(test_rows)}]")

        # Small delay to avoid overwhelming the API
        time.sleep(1)

    # Final save
    wb.save(EXCEL_PATH)

    print("\n" + "=" * 70)
    print(f"DONE: {total} tests | {success} success | {errors} errors")
    print(f"Results saved to: {EXCEL_PATH}")
    print("=" * 70)


if __name__ == "__main__":
    main()
