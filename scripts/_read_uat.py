import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\SSLTP11340\Desktop\SKYE\SKYE2.0 Feedback-UAT_Apr3.xlsx', data_only=True)
ws = wb['Feedback Sheet']
seen = set()
for r in range(2, ws.max_row+1):
    q = ws.cell(r,5).value
    score = ws.cell(r,7).value
    country = ws.cell(r,4).value
    tester = ws.cell(r,3).value
    accuracy = str(ws.cell(r,8).value or '')[:40]
    if q and q not in seen:
        seen.add(q)
        try:
            print(f'  [{score}] {country} | {tester} | {q} | {accuracy}')
        except UnicodeEncodeError:
            print(f'  [{score}] {country} | {tester} | {q.encode("ascii","replace").decode()} | {accuracy}')
