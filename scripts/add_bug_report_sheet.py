"""
add_bug_report_sheet.py
Adds a Bug Report sheet and a Summary sheet to SKYE_HR_Agent_Test_Cases.xlsx
"""

from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

EXCEL_PATH = r"c:\Users\SSLTP11340\Desktop\SKYE\Testing Prompts\SKYE_HR_Agent_Test_Cases.xlsx"

# ── Shared styles ────────────────────────────────────────────────────────────
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

def hdr(ws, row, col, value, bg="2F5496", fg="FFFFFF", bold=True, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=11, color=fg)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = CENTER if center else WRAP
    c.border = THIN
    return c

def cell(ws, row, col, value, wrap=True, bold=False, color=None, bg=None, center=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=10, bold=bold, color=color or "000000")
    if bg:
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = CENTER if center else WRAP
    c.border = THIN
    return c


# ════════════════════════════════════════════════════════════════════════════
# BUG DATA
# ════════════════════════════════════════════════════════════════════════════
BUGS = [
    {
        "id": 1, "sno": 34, "severity": "High",
        "category": "Cross-Country RBAC",
        "title": "US employee got India payroll policy",
        "user_email": "marcus.sternberg@hitachids.com",
        "user_country": "US",
        "question": "What is the payroll cycle in India?",
        "expected": "DENIED — US employee cannot access India payroll",
        "actual": "System returned full India payroll cycle details",
        "root_cause": "Geo-fence gap — country scope not enforced for US→India pair",
        "recommendation": "Review access_control_agent scope filtering for cross-country queries",
    },
    {
        "id": 2, "sno": 36, "severity": "High",
        "category": "Cross-Country RBAC",
        "title": "Poland employee got US benefits policy",
        "user_email": "bartek.marczak@hitachids.com",
        "user_country": "Poland",
        "question": "Where can I find my US benefits?",
        "expected": "DENIED — Poland employee cannot access US benefits",
        "actual": "System returned full US benefits info with links",
        "root_cause": "Geo-fence gap — country scope not enforced for Poland→US pair",
        "recommendation": "Consistent geo-fence enforcement across all country pairs",
    },
    {
        "id": 3, "sno": 42, "severity": "High",
        "category": "Manager Geo-Fence",
        "title": "Manager accessed Japan policy (no reportees in Japan)",
        "user_email": "vinay.rao@hitachidigital.com",
        "user_country": "India (reportees: IN, US, GB, AR)",
        "question": "What is the leave policy in Japan?",
        "expected": "DENIED — vinay.rao has no reportees in Japan",
        "actual": "System returned Japan leave policy details",
        "root_cause": "vinay.rao's reportee country list being over-fetched from BigQuery — expanding beyond actual hierarchy",
        "recommendation": "Audit BigQuery query in access_control_agent for manager reportee country resolution",
    },
    {
        "id": 4, "sno": 43, "severity": "High",
        "category": "Manager Geo-Fence",
        "title": "Manager accessed Poland policy (no reportees in Poland)",
        "user_email": "vinay.rao@hitachidigital.com",
        "user_country": "India (reportees: IN, US, GB, AR)",
        "question": "What are the standard working hours in Poland?",
        "expected": "DENIED — vinay.rao has no reportees in Poland",
        "actual": "System returned Poland working hours",
        "root_cause": "Same as Bug #3 — vinay.rao BQ reportee over-fetch",
        "recommendation": "Same as Bug #3",
    },
    {
        "id": 5, "sno": 77, "severity": "High",
        "category": "P-Card Access Control",
        "title": "Manager got P-Card content (VP+ only)",
        "user_email": "vinay.rao@hitachidigital.com",
        "user_country": "India",
        "question": "Can I use the P-Card for travel expenses?",
        "expected": "DENIED — Managers cannot access P-Card policies (VP+ only)",
        "actual": "System returned P-Card travel usage rules",
        "root_cause": "vinay.rao may be incorrectly resolved as VP-level in role hierarchy; or P-Card role check not applied for his specific role",
        "recommendation": "Verify vinay.rao's job level in BQ; check pcard_orchestrator role gate logic",
    },
    {
        "id": 6, "sno": 84, "severity": "High",
        "category": "Employee Lookup",
        "title": "Manager looked up employee outside his reportee tree",
        "user_email": "vinay.rao@hitachidigital.com",
        "user_country": "India",
        "question": "What are the holidays for Kevin Zhao?",
        "expected": "DENIED — Kevin Zhao is a reportee of blair.bakr, NOT vinay.rao",
        "actual": "System returned holiday info for Kevin Zhao",
        "root_cause": "vinay.rao's BQ reportee resolution returning employees from other manager trees (over-fetch)",
        "recommendation": "Fix reportee hierarchy BQ query scope — same root fix as Bugs #3, #4, #5",
    },
    {
        "id": 7, "sno": 61, "severity": "Medium",
        "category": "Super Admin Bypass",
        "title": "Super Admin employee search was denied",
        "user_email": "imtiaz.shaikh@hitachidigital.com",
        "user_country": "All (bypass)",
        "question": "Find employee Marcus Sternberg",
        "expected": "Super Admin bypass — should return employee details",
        "actual": "System returned 'do not have permission' / access denied",
        "root_cause": "Employee lookup feature may not be included in Super Admin bypass scope in access_control_agent",
        "recommendation": "Extend Super Admin bypass to cover employee search/lookup queries",
    },
    {
        "id": 8, "sno": 118, "severity": "Medium",
        "category": "Knowledge Base Gap",
        "title": "VPF opt-in query returned denial instead of policy",
        "user_email": "kamireddy.krishnareddy@hitachids.com",
        "user_country": "India",
        "question": "How to opt in for VPF?",
        "expected": "Should return VPF registration/opt-in process",
        "actual": "System denied / returned no information",
        "root_cause": "VPF (Voluntary Provident Fund) content may not be indexed in the knowledge base, or chunk retrieval failed",
        "recommendation": "Check if VPF policy doc is ingested; add 'VPF', 'voluntary provident fund' as retrieval keywords",
    },
    {
        "id": 9, "sno": 137, "severity": "Low",
        "category": "Knowledge Base Gap",
        "title": "GlobalLogic leave policy query returned denial",
        "user_email": "marcus.sternberg@hitachids.com",
        "user_country": "US",
        "question": "What is the GlobalLogic leave policy?",
        "expected": "Should return GL-labeled leave info",
        "actual": "System returned denial / no information",
        "root_cause": "GlobalLogic (GL) leave policies may be scoped differently, use a different OPCO tag, or not indexed under 'GlobalLogic leave policy'",
        "recommendation": "Verify GL leave docs are ingested with correct OPCO tag; test with 'GL' as opco filter",
    },
]

SEV_COLOR = {"High": "FFC7CE", "Medium": "FFEB9C", "Low": "C6EFCE"}
SEV_FONT  = {"High": "9C0006", "Medium": "9C6500", "Low": "006100"}


def build_bug_report_sheet(wb):
    # Remove existing if re-running
    if "Bug Report" in wb.sheetnames:
        del wb["Bug Report"]

    ws = wb.create_sheet("Bug Report")

    # ── Title banner ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    t = ws.cell(row=1, column=1, value="SKYE HR Agent — Bug Report  |  April 2026  |  Role-Based Access Testing")
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        ("Bug #", 6), ("S.No\n(Test Sheet)", 9), ("Severity", 11),
        ("Category", 20), ("Bug Title", 35), ("User Email", 35),
        ("User Country", 18), ("Question Asked", 40),
        ("Expected Behavior", 38), ("Actual Behavior", 38),
        ("Root Cause", 42), ("Recommendation", 42),
    ]
    for col_idx, (h, w) in enumerate(headers, 1):
        hdr(ws, 2, col_idx, h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[2].height = 30

    # ── Bug rows ──────────────────────────────────────────────────────────────
    for r, bug in enumerate(BUGS, 3):
        sev = bug["severity"]
        bg = SEV_COLOR[sev]
        fg = SEV_FONT[sev]

        cell(ws, r, 1, bug["id"], center=True, bold=True)
        cell(ws, r, 2, bug["sno"], center=True)
        c = ws.cell(row=r, column=3, value=sev)
        c.font = Font(name="Calibri", size=10, bold=True, color=fg)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = CENTER
        c.border = THIN
        cell(ws, r, 4, bug["category"])
        cell(ws, r, 5, bug["title"], bold=True)
        cell(ws, r, 6, bug["user_email"])
        cell(ws, r, 7, bug["user_country"])
        cell(ws, r, 8, bug["question"])
        cell(ws, r, 9, bug["expected"])
        cell(ws, r, 10, bug["actual"], bg="FFF2CC")
        cell(ws, r, 11, bug["root_cause"])
        cell(ws, r, 12, bug["recommendation"], bg="EBF5FB")
        ws.row_dimensions[r].height = 75

    ws.freeze_panes = "A3"

    # ── Summary box below bugs ────────────────────────────────────────────────
    summary_row = len(BUGS) + 4
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=12)
    s = ws.cell(row=summary_row, column=1,
                value="SUMMARY:  9 bugs found  |  High: 6  |  Medium: 2  |  Low: 1  |  Root cause pattern: vinay.rao BQ over-fetch affects 4/9 bugs (Bugs #3–6)")
    s.font = Font(name="Calibri", bold=True, size=11, color="1F3864")
    s.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    s.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    s.border = THIN
    ws.row_dimensions[summary_row].height = 22


def build_summary_sheet(wb):
    if "Test Summary" in wb.sheetnames:
        del wb["Test Summary"]

    ws = wb.create_sheet("Test Summary", 0)  # Insert as first sheet

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value="SKYE HR Agent — Test Execution Summary  |  April 15, 2026")
    t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    col_widths = [28, 14, 14, 14, 14, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Overall stats ─────────────────────────────────────────────────────────
    hdr(ws, 3, 1, "Metric", bg="2F5496")
    hdr(ws, 3, 2, "Count", bg="2F5496")
    hdr(ws, 3, 3, "%", bg="2F5496")
    hdr(ws, 3, 4, "", bg="2F5496")
    hdr(ws, 3, 5, "", bg="2F5496")
    hdr(ws, 3, 6, "Notes", bg="2F5496")

    ws.merge_cells("A2:F2")
    sub = ws.cell(row=2, column=1, value="Overall Test Results")
    sub.font = Font(name="Calibri", bold=True, size=12, color="1F3864")
    sub.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sub.alignment = CENTER

    stats = [
        ("Total Test Cases", 137, "100%", "", "", "137 cases across 15 scenario sections"),
        ("✅  Pass", 121, "88%", "", "", "Correct response returned as expected"),
        ("⚠️  Partial", 7,   "5%",  "", "", "Response present but language detection or content thin"),
        ("❌  Fail", 9,    "6%",  "", "", "Wrong behavior: data leak, incorrect denial, or missing content"),
    ]
    bg_map = {"✅  Pass": "C6EFCE", "⚠️  Partial": "FFEB9C", "❌  Fail": "FFC7CE"}
    fg_map = {"✅  Pass": "006100", "⚠️  Partial": "9C6500", "❌  Fail": "9C0006"}

    for r, (label, count, pct, _, __, note) in enumerate(stats, 4):
        bg = bg_map.get(label)
        fg = fg_map.get(label)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        cell(ws, r, 1, label, bold=(bg is not None), color=fg, bg=bg)
        cell(ws, r, 2, count, center=True, bold=(bg is not None), color=fg, bg=bg)
        cell(ws, r, 3, pct, center=True, color=fg, bg=bg)
        cell(ws, r, 6, note)
        ws.row_dimensions[r].height = 18

    # ── Section breakdown ─────────────────────────────────────────────────────
    section_row = 9
    ws.merge_cells(f"A{section_row}:F{section_row}")
    sub2 = ws.cell(row=section_row, column=1, value="Pass Rate by Scenario Section")
    sub2.font = Font(name="Calibri", bold=True, size=12, color="1F3864")
    sub2.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sub2.alignment = CENTER

    hdr(ws, section_row+1, 1, "Section")
    hdr(ws, section_row+1, 2, "Total")
    hdr(ws, section_row+1, 3, "Pass")
    hdr(ws, section_row+1, 4, "Partial")
    hdr(ws, section_row+1, 5, "Fail")
    hdr(ws, section_row+1, 6, "Notes / Key Findings")
    ws.row_dimensions[section_row+1].height = 20

    sections = [
        ("1. RBAC — Regular Employee",       31, 29, 2, 0,  "Strong — own-country access works well across India, US, Poland, Japan, France"),
        ("2. Cross-Country Denial",           6,  4,  0, 2,  "❌ US→India and Poland→US leaked; UK→India and India→UK correctly denied"),
        ("3. Manager Access",                 12, 9,  0, 3,  "❌ vinay.rao accessed Japan & Poland (no reportees); Germany/Sweden/Canada OK"),
        ("4. VP/Executive Access",            7,  6,  0, 1,  "❌ Super Admin employee search denied (Bug #7); rest passed"),
        ("5. Super Admin",                    4,  3,  0, 1,  "❌ Employee lookup blocked for Super Admin"),
        ("6. P-Card — Authorized",            12, 12, 0, 0,  "✅ All 12 VP/Admin P-Card questions answered correctly"),
        ("7. P-Card — Denied",                5,  4,  0, 1,  "❌ vinay.rao (Manager) got P-Card content — should be VP+ only"),
        ("8. Employee Lookup",                6,  4,  0, 2,  "❌ vinay.rao accessed non-reportee Kevin Zhao; regular emp correctly denied"),
        ("9. Multilingual — Auto-Detect",     12, 10, 2, 0,  "⚠️ Japanese/Chinese confusion (hiragana fix applied); French OK; Portuguese≈Spanish"),
        ("10. Explicit Translation",          4,  4,  0, 0,  "✅ All 4 passed — Hindi, French, Japanese, German"),
        ("11. T&E Policy",                    5,  5,  0, 0,  "✅ All 5 passed"),
        ("12. Leave / Payroll / Benefits",    16, 13, 2, 1,  "⚠️ VPF opt-in denied (Bug #8); WFH/Name-change responses thin"),
        ("13. Greetings & Small Talk",        4,  4,  0, 0,  "✅ All 4 passed — Hello, Thanks, Who is SKYE, AI chat"),
        ("14. Edge Cases & Fallback",         9,  9,  0, 0,  "✅ All 9 passed — shoes fallback, project contact, ambiguous, Hitachi Vantara"),
        ("15. OPCO Scenarios",               3,  2,  0, 1,  "❌ GlobalLogic leave policy denied (Bug #9); HD/HDS passed"),
    ]

    for r, (sec, tot, pas, par, fai, note) in enumerate(sections, section_row+2):
        cell(ws, r, 1, sec)
        cell(ws, r, 2, tot, center=True)
        cell(ws, r, 3, pas,  center=True, color="006100", bg="C6EFCE" if pas == tot else None)
        cell(ws, r, 4, par,  center=True, color="9C6500" if par else "000000", bg="FFEB9C" if par else None)
        cell(ws, r, 5, fai,  center=True, color="9C0006" if fai else "000000", bg="FFC7CE" if fai else None)
        cell(ws, r, 6, note)
        ws.row_dimensions[r].height = 20

    # ── Key findings ──────────────────────────────────────────────────────────
    kf_row = section_row + 2 + len(sections) + 1
    ws.merge_cells(f"A{kf_row}:F{kf_row}")
    kf = ws.cell(row=kf_row, column=1, value="Key Findings & Next Steps")
    kf.font = Font(name="Calibri", bold=True, size=12, color="1F3864")
    kf.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    kf.alignment = CENTER

    findings = [
        ("🔴 Priority 1", "Fix vinay.rao BQ over-fetch",
         "Bugs #3,4,5,6 all stem from vinay.rao's reportee tree returning too many countries/employees. Audit the manager hierarchy BQ query in access_control_agent.py."),
        ("🔴 Priority 2", "Fix cross-country geo-fence inconsistency",
         "Bugs #1,2: US→India and Poland→US deny not enforced. Check if country-code comparison is case-sensitive or uses different field names for these pairs."),
        ("🟡 Priority 3", "Extend Super Admin bypass to employee search",
         "Bug #7: Super Admins should be able to look up any employee. Update the bypass logic in access_control_agent.py to include employee lookup queries."),
        ("🟡 Priority 4", "Index missing KB content",
         "Bug #8 (VPF opt-in) and Bug #9 (GlobalLogic leave) — content either missing or not retrievable. Run ingestion check for these docs."),
    ]

    hdr(ws, kf_row+1, 1, "Priority")
    hdr(ws, kf_row+1, 2, "Area")
    ws.merge_cells(start_row=kf_row+1, start_column=3, end_row=kf_row+1, end_column=6)
    hdr(ws, kf_row+1, 3, "Action Required")

    for r, (pri, area, action) in enumerate(findings, kf_row+2):
        cell(ws, r, 1, pri, bold=True)
        cell(ws, r, 2, area, bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        cell(ws, r, 3, action)
        ws.row_dimensions[r].height = 45

    ws.freeze_panes = "A3"


# ── MAIN ─────────────────────────────────────────────────────────────────────
wb = load_workbook(EXCEL_PATH)
build_summary_sheet(wb)
build_bug_report_sheet(wb)

try:
    wb.save(EXCEL_PATH)
    print(f"Saved: {EXCEL_PATH}")
except PermissionError:
    alt = EXCEL_PATH.replace(".xlsx", "_with_bugs.xlsx")
    wb.save(alt)
    print(f"[!] Original locked — saved to: {alt}")
