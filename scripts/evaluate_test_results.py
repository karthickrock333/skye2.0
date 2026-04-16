"""
evaluate_test_results.py
Reads SKYE_HR_Agent_Test_Cases.xlsx, evaluates each response against
expected behavior, fills Pass/Fail (col 10) and Notes (col 11).
"""

import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXCEL_PATH = r"c:\Users\SSLTP11340\Desktop\SKYE\Testing Prompts\SKYE_HR_Agent_Test_Cases.xlsx"

# Column indices (1-based)
COL_SNO = 1
COL_CATEGORY = 2
COL_SUB = 3
COL_ROLE = 4
COL_EMAIL = 5
COL_COUNTRY = 6
COL_QUESTION = 7
COL_EXPECTED = 8
COL_RESPONSE = 9
COL_PASS_FAIL = 10
COL_NOTES = 11

PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PASS_FONT = Font(color="006100", bold=True)
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAIL_FONT = Font(color="9C0006", bold=True)
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
PARTIAL_FONT = Font(color="9C6500", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

# ── Denial keywords ──────────────────────────────────────────────────────────
DENIAL_PHRASES = [
    "do not have permission",
    "don't have permission",
    "not authorized",
    "access denied",
    "not have access",
    "you are not allowed",
    "cannot access",
    "restricted",
    "not available for your role",
    "not eligible",
    "permission to access",
    "don't have specific information",
    "i don't have specific information",
    "can only provide information applicable to",
    "raise a ticket",
    "please raise a ticket",
]

# Phrases that indicate the system returned actual policy content
CONTENT_INDICATORS = [
    "the policy", "the leave policy", "the payroll", "working hours",
    "maternity leave", "parental leave", "bereavement", "benefits",
    "401(k)", "fidelity", "cashless", "reimbursable", "eligible",
    "anti-bribery", "conflict of interest", "p-card", "pcard",
    "purchasing card", "transaction limit", "allowable",
    "resignation", "transfer", "salary", "loan", "insurance",
    "hospitalization", "creche", "pto", "fpto", "leave of absence",
    "annual leave", "casual leave", "vacation", "holiday",
    "sabbatical", "vpf", "nps", "flexi", "flexible benefit",
    "bank details", "bank account", "workday", "hinext",
    "business card", "expense report", "travel", "t&e",
    "id card", "employee id", "accommodation", "employment verification",
    "new hire", "onboarding", "day 1", "day-1", "visa",
    "conference", "scheduling", "resource management",
    "welcome", "skye", "hr assistant", "how can i help",
]

GREETING_PHRASES = [
    "hello", "hi ", "hey", "welcome", "how can i help",
    "skye", "hr assistant", "assist you", "glad to help",
]

FALLBACK_PHRASES = [
    "i'm not able to", "i am not able to", "i cannot help",
    "not an hr", "outside the scope", "not related to hr",
    "can't assist with that", "cannot assist",
    "raise a ticket", "askhr", "asknow",
]


def response_lower(resp):
    return (resp or "").lower().strip()


def has_any(text, phrases):
    return any(p in text for p in phrases)


def looks_like_denial(resp_low):
    """Check if response looks like a denial / no-info."""
    denial_signals = [
        "don't have specific information",
        "do not have permission",
        "don't have permission",
        "not authorized",
        "access denied",
        "cannot access",
        "not have access",
        "can only provide information applicable to",
        "not available for your role",
    ]
    return any(s in resp_low for s in denial_signals)


def looks_like_content(resp_low):
    """Check if response has substantial policy content (>100 chars and content keywords)."""
    if len(resp_low) < 80:
        return False
    return has_any(resp_low, CONTENT_INDICATORS)


def detect_language(text):
    """Language detection based on character ranges and vocabulary keywords."""
    if not text:
        return "unknown"

    # Japanese: hiragana or katakana MUST be checked before generic CJK
    if re.search(r'[\u3040-\u309f\u30a0-\u30ff]', text):
        return "japanese"
    # Hindi (Devanagari)
    if re.search(r'[\u0900-\u097f]', text):
        return "hindi"
    # Tamil
    if re.search(r'[\u0b80-\u0bff]', text):
        return "tamil"
    # Korean
    if re.search(r'[\u1100-\u11ff\uac00-\ud7af]', text):
        return "korean"
    # Chinese CJK (only if no hiragana/katakana found above)
    if re.search(r'[\u4e00-\u9fff]', text):
        return "chinese"

    # Vietnamese: only truly unique stacked-accent chars + đ (d-stroke).
    # Do NOT include é/è/ê — those also appear in French/Portuguese.
    if re.search(r'[đĐắằẵẳặấầẩẫậổộủứừ]', text):
        return "vietnamese"

    # European languages — vocabulary keywords (check before generic é/è fallback)
    low = text.lower()
    if any(w in low for w in ["arbeitszeit", "standardarbeitszeit", "urlaub", "stunden",
                               "mitarbeiter", "tage", "regelung", "anspruch"]):
        return "german"
    if any(w in low for w in ["urlop", "pracownik", "godzin", "godziny", "pracy",
                               "zatrudnieni", "miesięcy", "tygodni"]):
        return "polish"
    if any(w in low for w in ["congé", "congés", "politique", "employé", "jours ouvrables",
                               "travail", "semaines", "bénéficiez"]):
        return "french"
    if any(w in low for w in ["vacaciones", "política", "empleado", "días", "licencia",
                               "trabajo", "semanas", "permisos"]):
        return "spanish"
    if any(w in low for w in ["licença", "férias", "empregado", "política de licença",
                               "semanas", "trabalhadores", "remunerado"]):
        return "portuguese"
    if any(w in low for w in ["semester", "arbetstid", "anställd", "dagar", "ledighet",
                               "veckor", "föräldraledighet"]):
        return "swedish"
    return "english"


def evaluate_row(category, sub, role, email, country, question, expected, response):
    """
    Returns (verdict, note) where verdict is 'Pass', 'Fail', or 'Partial'.
    """
    if not response or not response.strip():
        return "Fail", "No response received"

    resp_low = response_lower(response)
    exp_low = (expected or "").lower().strip()
    q_low = (question or "").lower().strip()
    cat_low = (category or "").lower()

    # ─── DENIAL SCENARIOS ──────────────────────────────────────────────
    if "denied" in exp_low or "denied" in sub.lower() if sub else False:
        # Expected a denial
        if looks_like_denial(resp_low):
            return "Pass", "Correctly denied access"
        elif looks_like_content(resp_low) and len(resp_low) > 200:
            # Gave actual content when should have denied
            return "Fail", "Should have DENIED but returned policy content"
        else:
            # Ambiguous — might be indirect denial
            if "raise a ticket" in resp_low or "please raise" in resp_low:
                return "Partial", "Redirected to ticket (soft denial) — not explicit denial"
            return "Fail", "Expected denial but got non-denial response"

    # ─── P-CARD AUTHORIZED ─────────────────────────────────────────────
    if "p-card" in cat_low and "authorized" in cat_low:
        pcard_keywords = ["p-card", "pcard", "purchasing card", "procurement card",
                          "transaction limit", "allowable", "non-allowable",
                          "receipt", "misuse", "spending"]
        if has_any(resp_low, pcard_keywords):
            return "Pass", "P-Card content returned for authorized user"
        elif looks_like_denial(resp_low):
            return "Fail", "P-Card access DENIED for authorized VP/Admin user"
        else:
            return "Partial", "Response did not clearly contain P-Card details"

    # ─── EMPLOYEE LOOKUP ───────────────────────────────────────────────
    if "employee lookup" in cat_low:
        if "denied" in exp_low:
            if looks_like_denial(resp_low):
                return "Pass", "Correctly denied employee lookup"
            return "Fail", "Should have denied employee lookup"
        # Expected to find employee info
        if len(resp_low) > 100 and looks_like_content(resp_low):
            return "Pass", "Employee lookup returned relevant info"
        return "Partial", "Employee lookup response was thin"

    # ─── MULTILINGUAL ──────────────────────────────────────────────────
    if "multilingual" in cat_low or "translation" in cat_low:
        # Determine expected language
        expected_lang = None
        for lang in ["german", "polish", "japanese", "vietnamese", "chinese",
                      "hindi", "spanish", "tamil", "swedish", "french", "portuguese"]:
            if lang in exp_low:
                expected_lang = lang
                break

        detected_lang = detect_language(response)

        if expected_lang and detected_lang == expected_lang:
            return "Pass", f"Responded in {expected_lang} as expected"
        elif expected_lang and detected_lang == "english":
            # Check if the response is still useful but in wrong language
            if looks_like_content(resp_low) and len(resp_low) > 150:
                return "Partial", f"Content OK but in English, expected {expected_lang}"
            return "Fail", f"Responded in English, expected {expected_lang}"
        elif expected_lang:
            return "Partial", f"Expected {expected_lang}, detected {detected_lang}"
        # Fallback for translation requests
        if "answer in" in q_low:
            if detected_lang != "english":
                return "Pass", f"Responded in {detected_lang}"
            return "Partial", "Explicit translation requested but got English"
        return "Pass", "Multilingual response"

    # ─── GREETINGS ─────────────────────────────────────────────────────
    if "greeting" in cat_low or "greetings" in cat_low:
        if has_any(resp_low, GREETING_PHRASES) or len(resp_low) > 20:
            return "Pass", "Greeting handled appropriately"
        return "Fail", "Poor greeting response"

    # ─── EDGE CASES ────────────────────────────────────────────────────
    if "edge case" in cat_low:
        if "fallback" in exp_low or "not an hr" in exp_low:
            if has_any(resp_low, FALLBACK_PHRASES) or "not" in resp_low[:100]:
                return "Pass", "Correctly fell back for non-HR question"
            if looks_like_content(resp_low):
                return "Partial", "Attempted to answer non-HR question"
            return "Partial", "Fallback response unclear"
        if "clarification" in exp_low:
            if "?" in response or "which" in resp_low or "specify" in resp_low or "could you" in resp_low:
                return "Pass", "Asked for clarification as expected"
            return "Partial", "Did not explicitly ask for clarification"
        # Generic edge case — just check we got a reasonable response
        if len(resp_low) > 80:
            return "Pass", "Edge case returned substantive response"
        return "Partial", "Short response for edge case"

    # ─── RBAC REGULAR / MANAGER / VP / SUPER ADMIN (positive scenarios) ─
    # Expected to return policy content
    if "should return" in exp_low:
        if looks_like_content(resp_low) and len(resp_low) > 100:
            return "Pass", "Policy content returned as expected"
        elif looks_like_denial(resp_low):
            return "Fail", "Access was DENIED when policy content was expected"
        elif "raise a ticket" in resp_low and len(resp_low) < 300:
            return "Partial", "Redirected to ticket instead of providing policy"
        elif len(resp_low) > 60:
            return "Partial", "Response present but may lack detail"
        else:
            return "Fail", "Insufficient response"

    # ─── GENERIC FALLBACK ──────────────────────────────────────────────
    if len(resp_low) > 80:
        return "Pass", "Substantive response received"
    return "Partial", "Response may need manual review"


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    stats = {"Pass": 0, "Fail": 0, "Partial": 0, "skipped": 0}

    for row_idx in range(2, ws.max_row + 1):
        sno = ws.cell(row=row_idx, column=COL_SNO).value
        # Skip section header rows (merged cells or no S.No)
        if sno is None or not str(sno).strip().isdigit():
            continue

        category = ws.cell(row=row_idx, column=COL_CATEGORY).value or ""
        sub = ws.cell(row=row_idx, column=COL_SUB).value or ""
        role = ws.cell(row=row_idx, column=COL_ROLE).value or ""
        email = ws.cell(row=row_idx, column=COL_EMAIL).value or ""
        country = ws.cell(row=row_idx, column=COL_COUNTRY).value or ""
        question = ws.cell(row=row_idx, column=COL_QUESTION).value or ""
        expected = ws.cell(row=row_idx, column=COL_EXPECTED).value or ""
        response = ws.cell(row=row_idx, column=COL_RESPONSE).value or ""

        if not response.strip():
            ws.cell(row=row_idx, column=COL_PASS_FAIL, value="N/A")
            ws.cell(row=row_idx, column=COL_NOTES, value="No response")
            stats["skipped"] += 1
            continue

        verdict, note = evaluate_row(category, sub, role, email, country,
                                     question, expected, response)

        pf_cell = ws.cell(row=row_idx, column=COL_PASS_FAIL, value=verdict)
        note_cell = ws.cell(row=row_idx, column=COL_NOTES, value=note)
        note_cell.alignment = WRAP
        pf_cell.alignment = Alignment(horizontal="center", vertical="top")

        if verdict == "Pass":
            pf_cell.fill = PASS_FILL
            pf_cell.font = PASS_FONT
            stats["Pass"] += 1
        elif verdict == "Fail":
            pf_cell.fill = FAIL_FILL
            pf_cell.font = FAIL_FONT
            stats["Fail"] += 1
        else:
            pf_cell.fill = PARTIAL_FILL
            pf_cell.font = PARTIAL_FONT
            stats["Partial"] += 1

        print(f"  [{sno:>3}] {verdict:<7} | {note[:60]}")

    try:
        wb.save(EXCEL_PATH)
    except PermissionError:
        alt = EXCEL_PATH.replace(".xlsx", "_evaluated.xlsx")
        wb.save(alt)
        print(f"\n  [!] Original file locked — saved to: {alt}")

    total = stats["Pass"] + stats["Fail"] + stats["Partial"]
    print(f"\n{'='*60}")
    print(f"EVALUATION COMPLETE: {total} test cases evaluated")
    print(f"  Pass:    {stats['Pass']:>3}  ({stats['Pass']*100//total}%)" if total else "")
    print(f"  Partial: {stats['Partial']:>3}  ({stats['Partial']*100//total}%)" if total else "")
    print(f"  Fail:    {stats['Fail']:>3}  ({stats['Fail']*100//total}%)" if total else "")
    if stats["skipped"]:
        print(f"  Skipped: {stats['skipped']:>3}")
    print(f"{'='*60}")
    print(f"Results saved to: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
