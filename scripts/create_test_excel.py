"""
create_test_excel.py  (v2)
Generates a structured Excel for testing the main Skye HR Agent.
- P-Card: VP/Executive/Super Admin/HR-Finance only
- Diverse policy questions (not just holidays) sourced from UAT feedback
- Multilingual questions
- Correct email-to-scenario alignment
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="1F3864")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

HEADERS = [
    "S.No", "Scenario Category", "Sub-Scenario", "User Role",
    "User Email", "User Country", "Question",
    "Expected Behavior", "Response", "Pass/Fail", "Notes"
]

COL_WIDTHS = [6, 22, 30, 16, 38, 20, 55, 45, 55, 10, 30]

ws = wb.active
ws.title = "Skye HR Agent - Test Cases"

for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = THIN_BORDER
ws.row_dimensions[1].height = 30
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

rows = []
sno = [0]

def section(category):
    rows.append(("SECTION", category))

def tc(category, sub, role, email, country, question, expected):
    sno[0] += 1
    rows.append((sno[0], category, sub, role, email, country, question, expected))


# ═════════════════════════════════════════════════════════════════════════════
# 1. RBAC — Regular Employee (own country + global) — DIVERSE POLICY Qs
# ═════════════════════════════════════════════════════════════════════════════
section("1. RBAC — Regular Employee (Own Country + Global)")

tc("RBAC - Regular Employee", "Leave policy - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What is my leave policy in India?",
   "Should return India leave policy summary (UAT score 5)")

tc("RBAC - Regular Employee", "Payroll cycle - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What is the payroll cycle in India?",
   "Should return India payout date info (UAT score 5)")

tc("RBAC - Regular Employee", "Loan application - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I want to apply for a loan",
   "Should return loan application process + KB link (UAT score 5)")

tc("RBAC - Regular Employee", "Salary advance - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "How many times I can avail salary advance in a FY?",
   "Should return salary advance frequency rules (UAT score 5)")

tc("RBAC - Regular Employee", "Cashless hospitalization - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I need to get admitted, please help me with the cashless hospitalization process",
   "Should return cashless hospitalization process details (UAT score 5)")

tc("RBAC - Regular Employee", "Bank details update - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "How to update bank details?",
   "Should explain bank account update process (UAT score 5)")

tc("RBAC - Regular Employee", "Tax deductions SPOC - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Any SPOC whom I can reach out to for tax deductions?",
   "Should return tax deduction contact/process (UAT score 5)")

tc("RBAC - Regular Employee", "Creche facility - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I need to avail the creche facility",
   "Should return creche facility details (UAT score 4)")

tc("RBAC - Regular Employee", "Health insurance parents - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I want health insurance for my parents",
   "Should return parent insurance enrollment info")

tc("RBAC - Regular Employee", "Resignation process - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "How do I submit resignation?",
   "Should return resignation submission process (UAT score 5)")

tc("RBAC - Regular Employee", "Internal transfer - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "How can I track my internal transfer request?",
   "Should return internal transfer tracking info (UAT score 5)")

tc("RBAC - Regular Employee", "Salary query contact - India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I have some questions related to my salary, whom should I check with?",
   "Should return salary query escalation path (UAT score 5)")

tc("RBAC - Regular Employee", "Working hours - Poland",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "What are the standard working hours in Poland?",
   "Should return Poland working hours (UAT score 5)")

tc("RBAC - Regular Employee", "Salary range - Poland",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "Where can I find my salary range?",
   "Should return salary range lookup info (UAT score 4)")

tc("RBAC - Regular Employee", "Benefits - US",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "Where can I find my US benefits?",
   "Should return US benefits summary + link")

tc("RBAC - Regular Employee", "401k info - US",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "Where can I see my 401(k) info?",
   "Should return 401(k) info location + external site link")

tc("RBAC - Regular Employee", "Reasonable accommodation - US",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "How can I request a reasonable accommodation?",
   "Should return accommodation request process + policy link")

tc("RBAC - Regular Employee", "Employment verification - US",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "Who verifies my employment in the US?",
   "Should return US employment verification process")

tc("RBAC - Regular Employee", "Bereavement leave - US",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "How many days off can I take if my grandfather passed away in the US?",
   "Should return US bereavement leave policy")

tc("RBAC - Regular Employee", "FPTO during LOA - US",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "Can I use FPTO during a leave of absence?",
   "Should return FPTO usage rules during LOA")

tc("RBAC - Regular Employee", "Maternity leave - France",
   "Regular Employee", "nicola.lipreri@hitachids.com", "France",
   "What is the maternity leave policy in France?",
   "Should return France maternity leave policy")

tc("RBAC - Regular Employee", "Leave policy - Japan",
   "Regular Employee", "annapoorani.gopi@hitachids.com", "Japan",
   "What is the leave policy in Japan?",
   "Should return Japan leave policy details")

tc("RBAC - Regular Employee", "Wellness benefits - Japan",
   "Regular Employee", "annapoorani.gopi@hitachids.com", "Japan",
   "Are there any wellness benefits for Japan?",
   "Should return Japan wellness/health benefit info")

tc("RBAC - Regular Employee", "Submit a ticket (Global)",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "How do I submit a ticket?",
   "Should return ticket submission process + AskNow link (UAT score 5)")

tc("RBAC - Regular Employee", "Find employee ID (Global)",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "How do I find my employee ID?",
   "Should return employee ID lookup process + HiNext link (UAT score 5)")

tc("RBAC - Regular Employee", "Order business cards (Global)",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "How do I order business cards?",
   "Should return business card ordering process (UAT score 4)")

tc("RBAC - Regular Employee", "Anti-bribery policy (Global)",
   "Regular Employee", "dawid.stankiewicz@hitachids.com", "UK",
   "What is the anti-bribery policy?",
   "Should return global anti-bribery policy")

tc("RBAC - Regular Employee", "T&E reimbursable expenses (Global)",
   "Regular Employee", "saurav.kumar3@hitachids.com", "Germany",
   "What expenses are reimbursable under the T&E policy?",
   "Should return reimbursable expense categories")

tc("RBAC - Regular Employee", "Conflict of interest (Global)",
   "Regular Employee", "sang.phan@hitachids.com", "Vietnam",
   "What is the conflict of interest policy?",
   "Should return COI policy details")

tc("RBAC - Regular Employee", "Update personal details (Global)",
   "Regular Employee", "vimalraj.thanaraj@hitachids.com", "Singapore",
   "How can I update my personal details in the HR system?",
   "Should return HR system update process (UAT score 4)")

tc("RBAC - Regular Employee", "Workday password reset (Global)",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I forgot my Workday password. How can I reset it?",
   "Should return Workday password reset instructions (UAT score 4)")

tc("RBAC - Regular Employee", "Teams photo update (Global)",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I want to update photo on my Teams account, how can this be done?",
   "Should return Teams photo update process (UAT score 5)")

# ═════════════════════════════════════════════════════════════════════════════
# 2. RBAC — Cross-Country Denial
# ═════════════════════════════════════════════════════════════════════════════
section("2. RBAC — Cross-Country Denial")

tc("RBAC - Cross-Country Denial", "India emp asks UK parental leave",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What is the shared parental leave policy in UK?",
   "DENIED — India employee cannot access UK-specific policies")

tc("RBAC - Cross-Country Denial", "US emp asks India payroll",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "What is the payroll cycle in India?",
   "DENIED — US employee cannot access India payroll")

tc("RBAC - Cross-Country Denial", "Germany emp asks Japan benefits",
   "Regular Employee", "saurav.kumar3@hitachids.com", "Germany",
   "What are the wellness benefits in Japan?",
   "DENIED — Germany employee cannot access Japan policies")

tc("RBAC - Cross-Country Denial", "Poland emp asks US benefits",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "Where can I find my US benefits?",
   "DENIED — Poland employee cannot access US benefits")

tc("RBAC - Cross-Country Denial", "UK emp asks India leave",
   "Regular Employee", "dawid.stankiewicz@hitachids.com", "UK",
   "How many casual leaves are allowed per year in India?",
   "DENIED — UK employee cannot access India leave policy")

tc("RBAC - Cross-Country Denial", "India emp asks Poland maternity",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What is the maternity leave policy for Poland?",
   "DENIED — India employee cannot access Poland policies (tested in UAT)")

# ═════════════════════════════════════════════════════════════════════════════
# 3. RBAC — Manager Access
# ═════════════════════════════════════════════════════════════════════════════
section("3. RBAC — Manager Access (Own + Reportees' Countries)")

tc("RBAC - Manager", "Own country leave policy",
   "Manager", "vinay.rao@hitachidigital.com", "India (reports: IN,US,GB,AR)",
   "What is my leave policy in India?",
   "Should return India leave policy — own country")

tc("RBAC - Manager", "Reportee country bereavement (US)",
   "Manager", "vinay.rao@hitachidigital.com", "India (reports: IN,US,GB,AR)",
   "What is the bereavement leave policy in the US?",
   "Should return US bereavement policy — has reportees in US")

tc("RBAC - Manager", "Reportee country parental leave (UK)",
   "Manager", "vinay.rao@hitachidigital.com", "India (reports: IN,US,GB,AR)",
   "What is the shared parental leave policy in the UK?",
   "Should return UK parental leave — has reportees in GB")

tc("RBAC - Manager", "Non-reportee DENIED (Japan)",
   "Manager", "vinay.rao@hitachidigital.com", "India (reports: IN,US,GB,AR)",
   "What is the leave policy in Japan?",
   "DENIED — no reportees in Japan")

tc("RBAC - Manager", "Non-reportee DENIED (Poland)",
   "Manager", "vinay.rao@hitachidigital.com", "India (reports: IN,US,GB,AR)",
   "What are the standard working hours in Poland?",
   "DENIED — no reportees in Poland")

tc("RBAC - Manager", "Reportee country maternity (Canada)",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "What is the maternity leave policy for Canada?",
   "Should return Canada maternity leave — has reportees in CA")

tc("RBAC - Manager", "Reportee country hours (Poland)",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "What are the standard working hours in Poland?",
   "Should return Poland working hours — has reportees in PL")

tc("RBAC - Manager", "Reportee country payroll (India)",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "What is the payroll cycle in India?",
   "Should return India payroll cycle — has reportees in IN")

tc("RBAC - Manager", "Non-reportee DENIED (Germany)",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "What is the maternity leave policy in Germany?",
   "DENIED — no reportees in Germany")

tc("RBAC - Manager", "Reportee country hours (Germany)",
   "Manager", "rahul.umap@hitachids.com", "India (reports: IN,CN,PT,BR,SE,DE)",
   "What are the standard working hours in Germany?",
   "Should return Germany working hours — has reportees in DE")

tc("RBAC - Manager", "Reportee country leave (Sweden)",
   "Manager", "rahul.umap@hitachids.com", "India (reports: IN,CN,PT,BR,SE,DE)",
   "What is the leave policy in Sweden?",
   "Should return Sweden leave policy — has reportees in SE")

tc("RBAC - Manager", "New hire journey (Global)",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "How do I create a new hire journey as a manager?",
   "Should return manager onboarding journey process (UAT score 5)")

# ═════════════════════════════════════════════════════════════════════════════
# 4. RBAC — VP/Executive Access
# ═════════════════════════════════════════════════════════════════════════════
section("4. RBAC — VP/Executive Access (2-Level Reportees)")

tc("RBAC - VP/Executive", "VP own country benefits (US)",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "Where can I find my US benefits?",
   "Should return US benefits info — own country")

tc("RBAC - VP/Executive", "VP reportee country loan (India)",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "What is the loan policy in India?",
   "Should return India loan policy — has reportees in IN")

tc("RBAC - VP/Executive", "VP reportee country leave (Mexico)",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "What is the leave policy in Mexico?",
   "Should return Mexico leave — has reportee in MX")

tc("RBAC - VP/Executive", "VP reportee country hours (Poland)",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "What are the working hours in Poland?",
   "Should return Poland working hours — has reportee in PL")

tc("RBAC - VP/Executive", "VP India - payroll query",
   "VP", "kumar.ns@hitachids.com", "India",
   "What is the payroll cycle in India?",
   "Should return India payroll cycle — own country")

tc("RBAC - VP/Executive", "VP UK - parental leave",
   "VP", "duncan.mears@hitachids.com", "UK",
   "What is the shared parental leave policy?",
   "Should return UK shared parental leave — own country")

tc("RBAC - VP/Executive", "VP Finance - salary components",
   "VP/Finance", "rajat.mehta@hitachidigital.com", "India",
   "What are the salary components for employees?",
   "Should return salary components — VP Finance has elevated access")

# ═════════════════════════════════════════════════════════════════════════════
# 5. RBAC — Super Admin
# ═════════════════════════════════════════════════════════════════════════════
section("5. RBAC — Super Admin (Full Bypass)")

tc("RBAC - Super Admin", "Query any country (UK leave)",
   "Super Admin", "imtiaz.shaikh@hitachidigital.com", "All (bypass)",
   "What is the shared parental leave policy in UK?",
   "Should return UK policy — super admin bypass")

tc("RBAC - Super Admin", "Query any country (India payroll)",
   "Super Admin", "maria.luna@hitachidigital.com", "All (bypass)",
   "What is the payroll cycle in India?",
   "Should return India payroll — super admin bypass")

tc("RBAC - Super Admin", "Query any country (Poland hours)",
   "Super Admin", "shrimant.jaruhar@hitachidigital.com", "All (bypass)",
   "What are the standard working hours in Poland?",
   "Should return Poland working hours — super admin bypass")

tc("RBAC - Super Admin", "Global employee search",
   "Super Admin", "imtiaz.shaikh@hitachidigital.com", "All (bypass)",
   "Find employee Marcus Sternberg",
   "Should return employee details — super admin can search globally")

# ═════════════════════════════════════════════════════════════════════════════
# 6. P-CARD — AUTHORIZED (VP/Exec/Super Admin/HR-Finance ONLY)
# ═════════════════════════════════════════════════════════════════════════════
section("6. P-Card — AUTHORIZED (VP/Exec/Super Admin/HR-Finance)")

tc("P-Card - Authorized", "VP asks P-Card eligibility",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "Who is eligible to receive a P-Card?",
   "Should return P-Card eligibility criteria — VP is authorized")

tc("P-Card - Authorized", "VP asks allowable purchases",
   "VP", "kumar.ns@hitachids.com", "India",
   "What purchases are allowable on a P-Card?",
   "Should return allowable purchase list from gold-source PNG")

tc("P-Card - Authorized", "VP asks non-allowable purchases",
   "VP", "duncan.mears@hitachids.com", "UK",
   "What purchases are NOT allowable on P-Card?",
   "Should return non-allowable purchase list")

tc("P-Card - Authorized", "VP asks P-Card transaction limit",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "What is the transaction limit on the P-Card?",
   "Should return P-Card transaction/monthly limits")

tc("P-Card - Authorized", "VP asks P-Card for travel",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "Can I use the P-Card for travel and T&E expenses?",
   "Should clarify P-Card vs T&E card usage rules")

tc("P-Card - Authorized", "VP asks P-Card receipt requirements",
   "VP", "kumar.ns@hitachids.com", "India",
   "What are the receipt requirements for P-Card transactions?",
   "Should return P-Card receipt/documentation policy")

tc("P-Card - Authorized", "VP asks P-Card sharing",
   "VP", "duncan.mears@hitachids.com", "UK",
   "Can I share my P-Card with another employee?",
   "Should explain P-Card sharing policy (likely prohibited)")

tc("P-Card - Authorized", "VP asks P-Card personal use",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "What happens if I accidentally use the P-Card for personal expenses?",
   "Should explain personal use policy and rectification process")

tc("P-Card - Authorized", "VP asks P-Card for gifts",
   "VP", "kumar.ns@hitachids.com", "India",
   "Can I use the P-Card for employee gifts or recognition?",
   "Should return gift/recognition rules (may trigger footnote policies)")

tc("P-Card - Authorized", "VP asks P-Card misuse",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "What are the consequences of P-Card misuse?",
   "Should return P-Card misuse/fraud consequences")

tc("P-Card - Authorized", "Super Admin asks P-Card limit",
   "Super Admin", "imtiaz.shaikh@hitachidigital.com", "All (bypass)",
   "What is the P-Card monthly spending cap?",
   "Should return P-Card spending limits — super admin access")

tc("P-Card - Authorized", "VP Finance asks P-Card approval",
   "VP/Finance", "rajat.mehta@hitachidigital.com", "India",
   "What are the P-Card approval levels?",
   "Should return P-Card approval hierarchy — HR-Finance access")

# ═════════════════════════════════════════════════════════════════════════════
# 7. P-CARD — DENIED (Regular Employees / Managers)
# ═════════════════════════════════════════════════════════════════════════════
section("7. P-Card — DENIED (Regular Employees / Managers)")

tc("P-Card - DENIED", "Regular emp asks P-Card eligibility",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Who is eligible to receive a P-Card?",
   "DENIED — 'you do not have permission to access P-Card information'")

tc("P-Card - DENIED", "Regular emp asks allowable purchases",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "What purchases are allowable on a P-Card?",
   "DENIED — regular employee cannot access P-Card policies")

tc("P-Card - DENIED", "Regular emp asks P-Card limit",
   "Regular Employee", "dawid.stankiewicz@hitachids.com", "UK",
   "What is the P-Card transaction limit?",
   "DENIED — regular employee cannot access P-Card policies")

tc("P-Card - DENIED", "Manager asks P-Card for travel",
   "Manager", "vinay.rao@hitachidigital.com", "India (reports: IN,US,GB,AR)",
   "Can I use the P-Card for travel expenses?",
   "DENIED — managers (non-VP) cannot access P-Card policies")

tc("P-Card - DENIED", "Manager asks P-Card eligibility",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "Who is eligible for a P-Card?",
   "DENIED — managers (non-VP) cannot access P-Card policies")

# ═════════════════════════════════════════════════════════════════════════════
# 8. Employee Lookup
# ═════════════════════════════════════════════════════════════════════════════
section("8. Employee Lookup")

tc("Employee Lookup", "Manager looks up reportee (CN)",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "What is the leave policy for my employee Kevin Zhao?",
   "Should return CN leave policy for Kevin Zhao — direct reportee")

tc("Employee Lookup", "Manager looks up reportee (GB)",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "What is the leave policy for Richard Margetts?",
   "Should return UK leave policy — Richard is direct reportee in GB")

tc("Employee Lookup", "Manager holiday validation for reportee",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "Will my employee Kevin Zhao have a holiday on October 1st 2026?",
   "Should validate Oct 1 (China National Day) for Kevin Zhao in CN")

tc("Employee Lookup", "VP looks up direct reportee",
   "VP", "ben.fellows@hitachidigital.com", "US",
   "What are the vacations of my employee Sunil Kumar Sharma?",
   "Should return India vacation info — Sunil is direct reportee in IN")

tc("Employee Lookup", "Regular employee DENIED",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "What is the leave policy for Kamireddy Krishna Reddy?",
   "DENIED — regular employees cannot look up other employees")

tc("Employee Lookup", "Manager DENIED non-reportee",
   "Manager", "vinay.rao@hitachidigital.com", "India (reports: IN,US,GB,AR)",
   "What are the holidays for Kevin Zhao?",
   "DENIED — Kevin Zhao is NOT a reportee of vinay.rao")

# ═════════════════════════════════════════════════════════════════════════════
# 9. Multilingual — Auto-Detect
# ═════════════════════════════════════════════════════════════════════════════
section("9. Multilingual — Auto-Detect Language")

tc("Multilingual", "German - working hours",
   "Regular Employee", "saurav.kumar3@hitachids.com", "Germany",
   "Was sind die Standardarbeitszeiten in Deutschland?",
   "Detect German, respond in German with Germany working hours")

tc("Multilingual", "Polish - paid leave entitlement",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "Do ilu dni p\u0142atnego urlopu mam prawo w ci\u0105gu roku?",
   "Detect Polish, respond in Polish with paid leave info (UAT score 5)")

tc("Multilingual", "Polish - leave accrual increase",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "Mam na my\u015bli kiedy otrzymam wi\u0119cej dni urlopu z 20 na 26?",
   "Detect Polish, respond in Polish about leave accrual (UAT score 5)")

tc("Multilingual", "Japanese - leave policy",
   "Regular Employee", "annapoorani.gopi@hitachids.com", "Japan",
   "\u65e5\u672c\u306e\u4f11\u6687\u30dd\u30ea\u30b7\u30fc\u306f\u4f55\u3067\u3059\u304b\uff1f",
   "Detect Japanese, respond in Japanese with Japan leave info")

tc("Multilingual", "Vietnamese - leave policy",
   "Regular Employee", "sang.phan@hitachids.com", "Vietnam",
   "Ch\u00ednh s\u00e1ch ngh\u1ec9 ph\u00e9p \u1edf Vi\u1ec7t Nam l\u00e0 g\u00ec?",
   "Detect Vietnamese, respond in Vietnamese with Vietnam leave policy")

tc("Multilingual", "Chinese - payroll date",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "\u5de5\u8d44\u53d1\u653e\u65e5\u671f\u662f\u4ec0\u4e48\u65f6\u5019\uff1f",
   "Detect Chinese, respond in Chinese with payroll date info (UAT score 4)")

tc("Multilingual", "Hindi - leave application",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "\u092e\u0948\u0902 \u091b\u0941\u091f\u094d\u091f\u093f\u092f\u093e\u0901 \u0915\u0939\u093e\u0901 \u092a\u0947 \u092e\u093e\u0930\u094d\u0915 \u0915\u0930 \u0938\u0915\u0924\u0940 \u0939\u0942\u0901?",
   "Detect Hindi, respond in Hindi with leave application process")

tc("Multilingual", "Spanish - travel expense report",
   "Regular Employee", "victor.romero@hitachids.com", "Mexico",
   "\u00bfC\u00f3mo puedo presentar un informe de gastos de viaje?",
   "Detect Spanish, respond in Spanish with T&E report process")

tc("Multilingual", "Tamil - leave policy India",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "\u0b87\u0ba8\u0bcd\u0ba4\u0bbf\u0baf\u0bbe\u0bb5\u0bbf\u0bb2\u0bcd \u0bb5\u0bbf\u0b9f\u0bc1\u0bae\u0bc1\u0bb1\u0bc8 \u0ba8\u0bbe\u0b9f\u0bcd\u0b95\u0bb3\u0bcd \u0b8e\u0ba9\u0bcd\u0ba9?",
   "Detect Tamil, respond in Tamil with India leave/holiday info")

tc("Multilingual", "Swedish - leave policy",
   "Regular Employee", "johan.blomstrand@hitachids.com", "Sweden",
   "Vad \u00e4r semesterpolicyn i Sverige?",
   "Detect Swedish, respond in Swedish with Sweden leave info")

tc("Multilingual", "French - vacation policy",
   "Regular Employee", "nicola.lipreri@hitachids.com", "France",
   "Quelle est la politique de cong\u00e9s en France?",
   "Detect French, respond in French with France vacation policy")

tc("Multilingual", "Portuguese - leave policy",
   "Regular Employee", "josemiguel.mendes@hitachids.com", "Portugal",
   "Qual \u00e9 a pol\u00edtica de licen\u00e7as em Portugal?",
   "Detect Portuguese, respond in Portuguese with Portugal leave info")

section("10. Multilingual — Explicit Translation Request")

tc("Translation Request", "Answer in Hindi",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What is my leave policy in India? Answer in Hindi",
   "Respond in Hindi with India leave policy")

tc("Translation Request", "Answer in French",
   "Regular Employee", "nicola.lipreri@hitachids.com", "France",
   "What are the reimbursable expenses? Answer in French",
   "Respond in French with T&E reimbursable expenses")

tc("Translation Request", "Answer in Japanese",
   "Regular Employee", "annapoorani.gopi@hitachids.com", "Japan",
   "What is the leave policy? Answer in Japanese",
   "Respond in Japanese with Japan leave policy")

tc("Translation Request", "Manager query in German",
   "Manager", "rahul.umap@hitachids.com", "India (reports: IN,CN,PT,BR,SE,DE)",
   "Was sind die Standardarbeitszeiten in Deutschland?",
   "Respond in German with Germany working hours — has reportees in DE")

# ═════════════════════════════════════════════════════════════════════════════
# 11. T&E Policy
# ═════════════════════════════════════════════════════════════════════════════
section("11. Travel & Expense Policy")

tc("T&E Policy", "Expense report submission timeline",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What is the timeframe for submitting expense reports?",
   "Should return expense report submission deadline")

tc("T&E Policy", "Foreign currency handling",
   "Regular Employee", "saurav.kumar3@hitachids.com", "Germany",
   "How do I handle foreign currency expenses?",
   "Should return foreign currency expense process")

tc("T&E Policy", "Personal + business travel",
   "Regular Employee", "annapoorani.gopi@hitachids.com", "Japan",
   "Can I combine personal travel with a business trip?",
   "Should return rules for combining personal/business travel")

tc("T&E Policy", "Travel policy details",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Can you please tell my travel policy?",
   "Should return travel policy summary for India")

tc("T&E Policy", "Expense system to use",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Which expense system should I use to submit my expenses?",
   "Should return correct expense system info")

# ═════════════════════════════════════════════════════════════════════════════
# 12. Leave, Payroll & Benefits (Diverse)
# ═════════════════════════════════════════════════════════════════════════════
section("12. Leave, Payroll & Benefits (Diverse)")

tc("Leave/Benefits", "Cancel approved leave",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Can I cancel an approved leave?",
   "Should return leave cancellation process (UAT score 4)")

tc("Leave/Benefits", "Retroactive leave",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Can leave be applied retroactively? If yes, under what conditions?",
   "Should return retroactive leave rules (UAT score 4)")

tc("Leave/Benefits", "Casual leaves per year",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "How many casual leaves are allowed per year, and do they carry forward?",
   "Should return casual leave count + carry-forward rules (UAT score 4)")

tc("Leave/Benefits", "Emergency time off - LOP",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What happens if I exhaust all my earned leave but need emergency time off?",
   "Should return LOP / emergency leave rules (UAT score 4)")

tc("Leave/Benefits", "Sabbatical leave",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Is there any sabbatical leave type in the organization?",
   "Should explain sabbatical leave availability")

tc("Leave/Benefits", "VPF opt-in",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "How to opt in for VPF?",
   "Should return VPF registration/opt-in process")

tc("Leave/Benefits", "NPS enrollment",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I want to opt in for NPS please guide",
   "Should return NPS enrollment process")

tc("Leave/Benefits", "ID card replacement",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I lost my ID card, how do I get a new one?",
   "Should return ID card replacement process (UAT score 4)")

tc("Leave/Benefits", "Flexible benefits",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What are my flexi benefits?",
   "Should return flexible benefit plan details")

tc("Leave/Benefits", "Incorrect salary rectification",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I have received incorrect salary, I need support for rectification",
   "Should return salary discrepancy escalation process (UAT score 5)")

tc("Leave/Benefits", "Holiday on weekend",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "What happens if a holiday falls on a weekend?",
   "Should explain weekend holiday compensation policy")

tc("Leave/Benefits", "Carry forward unused holiday",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "Can I carry forward unused holiday entitlement to next year?",
   "Should explain holiday carry-forward rules for Poland")

tc("Leave/Benefits", "WFH policy",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Can an employee work from home permanently under company policy?",
   "Should return WFH/remote work policy")

tc("Leave/Benefits", "Name change after marriage",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Can I change my name in Workday after marriage?",
   "Should return name change process in Workday (UAT score 4)")

tc("Leave/Benefits", "Maternity leave (Canada) - Manager",
   "Manager", "blair.bakr@hitachidigital.com", "US (reports: CN,US,IN,CA,PL,SG,GB)",
   "What is the maternity leave policy for Canada?",
   "Should return Canada maternity leave — has reportees in CA")

tc("Leave/Benefits", "HR ticket link",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Can you share me link for raising HR ticket?",
   "Should return HR ticket submission link/process (UAT score 4)")

# ═════════════════════════════════════════════════════════════════════════════
# 13. Greetings & Small Talk
# ═════════════════════════════════════════════════════════════════════════════
section("13. Greetings & Small Talk")

tc("Greeting", "Hello",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "Hello",
   "Should return SKYE greeting message (UAT score 5)")

tc("Greeting", "Thanks",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "Thanks!",
   "Should return warm acknowledgement (UAT score 5)")

tc("Greeting", "Who is SKYE",
   "Regular Employee", "dawid.stankiewicz@hitachids.com", "UK",
   "Who are you / What is SKYE?",
   "Should introduce HD SKYE and capabilities (UAT score 5)")

tc("Greeting", "Do we have AI chat",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Do we have an AI chat function?",
   "Should confirm chat function availability + info")

# ═════════════════════════════════════════════════════════════════════════════
# 14. Edge Cases & Fallback
# ═════════════════════════════════════════════════════════════════════════════
section("14. Edge Cases & Fallback")

tc("Edge Case", "Non-HR question (shoes)",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "Where are my shoes?",
   "Should return fallback — not an HR question (UAT score 4)")

tc("Edge Case", "Not tagged to project",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I am not tagged to a project, whom do I contact?",
   "Should redirect to schedulingx@hitachidigital.com / Resource Management")

tc("Edge Case", "Poland payroll contact",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "Could you send me the payroll contact details for Poland?",
   "Should redirect to payroll.pl@hitachidigital.com")

tc("Edge Case", "Ambiguous question",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "Tell me about the policy",
   "Should ask for clarification or list available policy topics")

tc("Edge Case", "Hitachi Vantara NOT covered",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "What is the Hitachi Vantara leave policy?",
   "Should indicate HV policies not covered / redirect to AskNow")

tc("Edge Case", "New joiner bank details",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I recently joined and I missed providing my bank account details during onboarding. How can I do it now?",
   "Should return bank details update process for new joiners")

tc("Edge Case", "Day-1 documents Poland",
   "Regular Employee", "bartek.marczak@hitachids.com", "Poland",
   "What documents should I bring on Day 1 in Poland?",
   "Should return Poland Day-1 document checklist")

tc("Edge Case", "Visa support letter",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "I have a US visa that is going to expire. Can the company provide a support letter for renewal?",
   "Should return visa support/company letter process (UAT score 5)")

tc("Edge Case", "Onsite conference eligibility",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "Can I know whether I am eligible to make onsite conferences?",
   "Should return onsite conference eligibility rules")

# ═════════════════════════════════════════════════════════════════════════════
# 15. OPCO
# ═════════════════════════════════════════════════════════════════════════════
section("15. OPCO Scenarios")

tc("OPCO", "HD-specific T&E policy",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "What is the Hitachi Digital T&E policy?",
   "Should return HD-labeled T&E policy")

tc("OPCO", "HDS-specific payroll",
   "Regular Employee", "kamireddy.krishnareddy@hitachids.com", "India",
   "What are the HDS payroll policies?",
   "Should return HDS-labeled payroll information")

tc("OPCO", "GlobalLogic query",
   "Regular Employee", "marcus.sternberg@hitachids.com", "US",
   "What is the GlobalLogic leave policy?",
   "Should return GL-labeled leave info or indicate separate GL policies")


# ═══════════════════════════════════════════════════════════════════════════
# WRITE ALL ROWS
# ═══════════════════════════════════════════════════════════════════════════

current_row = 2
for item in rows:
    if item[0] == "SECTION":
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(HEADERS))
        cell = ws.cell(row=current_row, column=1, value=item[1])
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
        ws.row_dimensions[current_row].height = 24
    else:
        sno_val, cat, sub, role, email, country, question, expected = item
        values = [sno_val, cat, sub, role, email, country, question, expected, "", "", ""]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
        ws.row_dimensions[current_row].height = 45
    current_row += 1

output_path = r"c:\Users\SSLTP11340\Desktop\SKYE\Testing Prompts\SKYE_HR_Agent_Test_Cases.xlsx"
wb.save(output_path)
print(f"Excel created: {output_path}")
print(f"Total test cases: {sno[0]}")
print(f"Total rows (incl. sections): {current_row - 2}")
